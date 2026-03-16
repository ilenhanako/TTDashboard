"""
WHO DALY Global Categorisation Script
======================================
Analogue of categorise_daly.py, adapted for the multi-year global format
(ghe2021_daly_global_new.xlsx).

Key differences from the by-country script:
  - Input has one sheet per YEAR (not per age-group); age groups are COLUMNS
  - There are no countries — columns represent sex × age-group combinations
  - All 6 years (2000, 2010, 2015, 2019, 2020, 2021) are processed in one run
  - Extra output sheets show trends across years and the MPNc burden over time

Output workbook structure:
  ┌ Global 2021 ORIGINAL       Raw copy of the most-recent year (reference)
  ├ 2000 Categorised  ┐
  ├ 2010 Categorised  │  Selected GHE causes, rows = causes,
  ├ 2015 Categorised  │  columns = Both-sexes / Male / Female × age-groups
  ├ 2019 Categorised  │  (19 columns total per sheet)
  ├ 2020 Categorised  │
  ├ 2021 Categorised  ┘
  ├ Trends – Both Sexes Total  One column per year; both-sexes total only
  └ MPNc – Global              MPNc causes across all six years

Usage:
    python categorise_daly_global.py

Configure RAW_FILE and OUTPUT_FILE below.  No command-line arguments needed.
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


# ── CONFIGURATION ─────────────────────────────────────────────────────────────

RAW_FILE    = "GHE_DALYOriginal/global/ghe2021daly_global.xlsx"
OUTPUT_FILE = "GHE_DALYFinal/global/ghe2021_daly_global_categorised.xlsx"

# ── PATH RESOLUTION ──────────────────────────────────────────────────────────
# When the script is run alongside the raw file (same directory), these paths
# work as-is.  If running from a different working directory, update them to
# absolute paths, e.g.:
#   RAW_FILE    = "/path/to/ghe2021_daly_global_new.xlsx"
#   OUTPUT_FILE = "/path/to/output/ghe2021_daly_global_categorised.xlsx"

# Ordered list of (raw_sheet_name, year_string) pairs to process
YEAR_SHEETS = [
    ("Global 2000", "2000"),
    ("Global 2010", "2010"),
    ("Global 2015", "2015"),
    ("Global 2019", "2019"),
    ("Global 2020", "2020"),
    ("Global 2021", "2021"),
]

# The 19 sex × age-group value columns present in every Global YYYY sheet.
# In the raw DataFrame these map to 0-based column indices 6 – 24 (columns G – Y).
GLOBAL_COLUMNS = [
    ("Both sexes", "Total"),
    ("Male",       "Total"),
    ("Female",     "Total"),
    ("Male",       "0-28 days"),
    ("Male",       "1-59 months"),
    ("Male",       "5-14"),
    ("Male",       "15-29"),
    ("Male",       "30-49"),
    ("Male",       "50-59"),
    ("Male",       "60-69"),
    ("Male",       "70+"),
    ("Female",     "0-28 days"),
    ("Female",     "1-59 months"),
    ("Female",     "5-14"),
    ("Female",     "15-29"),
    ("Female",     "30-49"),
    ("Female",     "50-59"),
    ("Female",     "60-69"),
    ("Female",     "70+"),
]
N_COLS         = len(GLOBAL_COLUMNS)   # 19
DATA_COL_START = 6                     # 0-based index of first value column in raw DataFrame

# Number of fixed metadata columns before the value columns in every output row
# ColA=GHE code | ColB=col2 | ColC=section letter | ColD=section name | ColE=cause | ColF=spacer
FIXED_COLS = 6

NUMBER_FORMAT        = '#,##0.0'
PROGRESS_CHECKPOINTS = 8
MIN_PROGRESS_STRIDE  = 25


# ── PROGRESS UTILITIES (identical to categorise_daly.py) ─────────────────────

def get_progress_stride(total_rows, checkpoints=PROGRESS_CHECKPOINTS,
                        min_stride=MIN_PROGRESS_STRIDE):
    """Return a sensible progress-report interval for row-based loops."""
    if total_rows <= 0:
        return 1
    return max(1, min(total_rows, max(min_stride, total_rows // checkpoints)))


def print_progress(label, current, total):
    """Emit a consistent terminal progress message for long-running steps."""
    if total <= 0:
        print(f"    {label}: {current:,}")
        return
    pct = (current / total) * 100
    print(f"    {label}: {current:,}/{total:,} rows ({pct:5.1f}%)")


# ── STYLE PALETTE (identical to categorise_daly.py) ──────────────────────────

_STYLE_CACHE = None


def build_style_palette():
    """Centralised workbook style definitions (cached)."""
    global _STYLE_CACHE
    if _STYLE_CACHE is not None:
        return _STYLE_CACHE

    thin_gray   = Side(style='thin',   color='D9E2F3')
    medium_blue = Side(style='medium', color='5B9BD5')
    strong_navy = '1F4E78'

    palette = {
        'title': {
            'font':      Font(bold=True, size=12, color=strong_navy),
            'fill':      PatternFill('solid', fgColor='EAF3FF'),
            'alignment': Alignment(horizontal='left', vertical='center'),
        },
        'header': {
            'font':      Font(bold=True, color='FFFFFF'),
            'fill':      PatternFill('solid', fgColor=strong_navy),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border':    Border(left=medium_blue, right=medium_blue,
                                top=medium_blue, bottom=medium_blue),
        },
        'subheader': {
            'font':      Font(bold=True, color=strong_navy),
            'fill':      PatternFill('solid', fgColor='DCE6F1'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border':    Border(left=thin_gray, right=thin_gray,
                                top=thin_gray, bottom=thin_gray),
        },
        'section': {
            'font':      Font(bold=True, color=strong_navy),
            'fill':      PatternFill('solid', fgColor='D9EAF7'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border':    Border(left=thin_gray, right=thin_gray,
                                top=thin_gray, bottom=thin_gray),
        },
        'summary': {
            'font':      Font(bold=True, color='203864'),
            'fill':      PatternFill('solid', fgColor='EAF2F8'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border':    Border(left=thin_gray, right=thin_gray,
                                top=thin_gray, bottom=thin_gray),
        },
        'population': {
            'font':      Font(bold=True, color='7F6000'),
            'fill':      PatternFill('solid', fgColor='FFF2CC'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border':    Border(left=thin_gray, right=thin_gray,
                                top=thin_gray, bottom=thin_gray),
        },
        'calc': {
            'font':      Font(bold=True, italic=True, color='385723'),
            'fill':      PatternFill('solid', fgColor='E2F0D9'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border':    Border(left=thin_gray, right=thin_gray,
                                top=thin_gray, bottom=thin_gray),
        },
        'body': {
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border':    Border(left=thin_gray, right=thin_gray,
                                top=thin_gray, bottom=thin_gray),
        },
        'numeric': {
            'alignment': Alignment(horizontal='right', vertical='center'),
            'border':    Border(left=thin_gray, right=thin_gray,
                                top=thin_gray, bottom=thin_gray),
        },
        'blank': {
            'fill': PatternFill('solid', fgColor='F7F9FC'),
        },
        'band_fill': PatternFill('solid', fgColor='F8FBFF'),
    }
    _STYLE_CACHE = palette
    return palette


def apply_cell_style(cell, style_def=None, number_format=None):
    """Apply a partial style dict to a cell."""
    if style_def:
        if 'font'      in style_def: cell.font      = style_def['font']
        if 'fill'      in style_def: cell.fill      = style_def['fill']
        if 'alignment' in style_def: cell.alignment = style_def['alignment']
        if 'border'    in style_def: cell.border    = style_def['border']
    if number_format is not None:
        cell.number_format = number_format


# ── CALC HELPER FUNCTIONS (same logic as categorise_daly.py) ─────────────────

def _parasitic_excl(d):
    """Parasitic & vector diseases excl. Malaria (220) and Dengue (300)."""
    return d.get(210, 0) - d.get(220, 0) - d.get(300, 0)


def _cv_excl_stroke(d):
    """Cardiovascular diseases excl. Stroke (GHE 1100 – 1140)."""
    return d.get(1100, 0) - d.get(1140, 0)


def _neoplasms(d):
    """Neoplasms = Malignant neoplasms (610) + Other neoplasms (790)."""
    return d.get(610, 0) + d.get(790, 0)


def _mpnc_total(d):
    """Total MPNc = sum of all MPNc leaf codes (no double-counting)."""
    leaf = [420, 500, 510, 520, 530, 550, 560, 570, 580, 590, 1505]
    return sum(d.get(c, 0) for c in leaf)


# ── GHE LABEL MAP (same as categorise_daly.py) ───────────────────────────────

GHE_LABEL_MAP = {
    0:    (None, 'All Causes',   None, None),
    10:   (None, 'Communicable, maternal, perinatal and nutritional conditions', None, None),
    20:   (None, 'A.',  'Infectious and parasitic diseases', None),
    30:   (None, None,  None,   'Tuberculosis'),
    40:   (None, None,  None,   'STDs excluding HIV'),
    100:  (None, None,  None,   'HIV/AIDS'),
    110:  (None, None,  None,   'Diarrhoeal diseases'),
    120:  (None, None,  None,   'Childhood-cluster diseases'),
    170:  (None, None,  None,   'Meningitis'),
    180:  (None, None,  None,   'Encephalitis'),
    185:  (None, None,  None,   'Hepatitis'),
    210:  (None, None,  None,   'Parasitic and vector diseases'),
    220:  (None, None,  None,   'Malaria'),
    300:  (None, None,  None,   'Dengue'),
    330:  (None, None,  None,   'Intestinal nematode infections'),
    365:  (None, None,  None,   'Leprosy'),
    370:  (None, None,  None,   'Other infectious diseases'),
    380:  (None, None,  None,   'Respiratory Infectious'),
    420:  (None, None,  None,   'Maternal Conditions'),
    500:  (None, None,  None,   'Preterm birth complications'),
    510:  (None, None,  None,   'Birth asphyxia and birth trauma'),
    520:  (None, None,  None,   'Neonatal sepsis and infections'),
    530:  (None, None,  None,   'Other neonatal conditions'),
    550:  (None, None,  None,   'Protein-energy malnutrition'),
    560:  (None, None,  None,   'Iodine deficiency'),
    570:  (None, None,  None,   'Vitamin A deficiency'),
    580:  (None, None,  None,   'Iron-deficiency anaemia'),
    590:  (None, None,  None,   'Other nutritional deficiencies'),
    600:  (None, 'C.',  'Noncommunicable diseases', None),
    610:  (None, None,  None,   'Malignant neoplasms'),
    790:  (None, None,  None,   'Other neoplasms'),
    800:  (None, None,  None,   'Diabetes mellitus'),
    810:  (None, None,  None,   'Endocrine, blood, immune disorders'),
    830:  (None, None,  None,   'Depressive Disorders'),
    860:  (None, None,  None,   'Alcohol use disorders'),
    870:  (None, None,  None,   'Drug use disorders'),
    880:  (None, None,  None,   'Anxiety disorders'),
    890:  (None, None,  None,   'Eating disorders'),
    911:  (None, None,  None,   'Attention deficit/hyperactivity syndrome'),
    912:  (None, None,  None,   'Conduct disorder'),
    940:  (None, None,  None,   'Neurological conditions'),
    1020: (None, None,  None,   'Sense organ diseases'),
    1100: (None, None,  None,   'Cardiovascular diseases'),
    1140: (None, None,  None,   'Stroke'),
    1180: (None, None,  None,   'Chronic obstructive pulmonary disease'),
    1190: (None, None,  None,   'Asthma'),
    1200: (None, None,  None,   'Other respiratory diseases'),
    1210: (None, None,  None,   'Digestive diseases'),
    1260: (None, None,  None,   'Genitourinary diseases'),
    1330: (None, None,  None,   'Skin diseases'),
    1340: (None, None,  None,   'Musculoskeletal diseases'),
    1470: (None, None,  None,   'Oral conditions'),
    1505: (None, None,  None,   'Sudden infant death syndrome'),
    1510: (None, 'D.',  'Injuries incl War', None),
    1520: (None, None,  None,   'Unintentional injuries'),
    1610: (None, None,  None,   'Self-harm'),
    1620: (None, None,  None,   'Interpersonal violence'),
    1630: (None, None,  None,   'Collective violence and legal intervention'),
    1700: (None, None,  None,   'Other COVID-19 pandemic-related outcomes'),
}


# ── ROW TEMPLATE ──────────────────────────────────────────────────────────────

def _global_categorised_template():
    """
    Row template for global categorised sheets.

    Each entry is a tuple whose first element is the row 'kind':
      ('pop',)                           – global population row (from raw data)
      ('blank',)                         – empty spacer row
      ('header', col3, col4, col5)       – bold named section-header row
      ('data',   ghe_code)               – data row looked up from GHE_LABEL_MAP
      ('calc',   label, calc_fn)         – computed/derived row

    This selection mirrors _persons_all_ages_template() and _sex_section_template()
    from categorise_daly.py, expanded to include all three sex totals and all age
    groups via the column structure of the global file.
    """
    T = []
    T.append(('pop',))
    T.append(('blank',))
    T.append(('data', 0))    # All Causes
    T.append(('blank',))

    # ── Section A: Communicable, maternal, perinatal and nutritional ──────────
    T.append(('header', 'A.', 'Communicable, maternal, perinatal and nutritional conditions', None))
    T.append(('data', 10))   # Communicable total (summary)
    T.append(('data', 20))   # A. Infectious & parasitic (summary)
    T.append(('data', 30))   # Tuberculosis
    T.append(('data', 110))  # Diarrhoeal diseases
    T.append(('data', 185))  # Hepatitis
    T.append(('data', 220))  # Malaria
    T.append(('data', 300))  # Dengue
    T.append(('calc', 'Parasitic and vector diseases excl Malaria and Dengue', _parasitic_excl))
    T.append(('data', 370))  # Other infectious diseases
    T.append(('data', 380))  # Respiratory Infectious
    T.append(('data', 1700)) # Other COVID-19 pandemic-related outcomes
    T.append(('data', 170))  # Meningitis
    T.append(('data', 180))  # Encephalitis
    T.append(('data', 120))  # Childhood-cluster diseases
    T.append(('data', 330))  # Intestinal nematode infections
    T.append(('blank',))

    # ── Section B: Maternal, Perinatal and Nutritional ────────────────────────
    T.append(('header', 'B.', 'Maternal, Perinatal and Nutritional Conditions', None))
    T.append(('data', 420))  # Maternal conditions
    T.append(('data', 500))  # Preterm birth complications
    T.append(('data', 510))  # Birth asphyxia and birth trauma
    T.append(('data', 520))  # Neonatal sepsis and infections
    T.append(('data', 530))  # Other neonatal conditions
    T.append(('data', 550))  # Protein-energy malnutrition
    T.append(('data', 560))  # Iodine deficiency
    T.append(('data', 570))  # Vitamin A deficiency
    T.append(('data', 580))  # Iron-deficiency anaemia
    T.append(('data', 590))  # Other nutritional deficiencies
    T.append(('data', 1505)) # Sudden infant death syndrome
    T.append(('blank',))

    # ── Section C: Noncommunicable diseases ───────────────────────────────────
    T.append(('header', 'C.', 'Noncommunicable diseases', None))
    T.append(('data', 600))  # C. Noncommunicable diseases (summary)
    T.append(('data', 800))  # Diabetes mellitus
    T.append(('calc', 'Cardiovascular diseases excl Stroke', _cv_excl_stroke))
    T.append(('data', 1140)) # Stroke
    T.append(('data', 1180)) # Chronic obstructive pulmonary disease
    T.append(('data', 1190)) # Asthma
    T.append(('data', 1200)) # Other respiratory diseases
    T.append(('calc', 'Neoplasms', _neoplasms))
    T.append(('blank',))

    # ── Section D: Injuries ───────────────────────────────────────────────────
    T.append(('header', 'D.', 'Injuries incl War', None))
    T.append(('data', 1510)) # D. Injuries (summary)
    T.append(('data', 1520)) # Unintentional injuries
    T.append(('data', 1630)) # Collective violence and legal intervention
    T.append(('data', 1620)) # Interpersonal violence
    T.append(('blank',))

    # ── Section E: Mental Disorders & Wellbeing ───────────────────────────────
    T.append(('header', 'E.', 'Mental Disorders & Wellbeing', None))
    T.append(('data', 1610)) # Self-harm
    T.append(('data', 830))  # Depressive Disorders
    T.append(('data', 880))  # Anxiety disorders
    T.append(('data', 890))  # Eating disorders
    T.append(('data', 911))  # Attention deficit/hyperactivity syndrome
    T.append(('data', 912))  # Conduct disorder
    T.append(('blank',))

    # ── Section F: Substance Abuse ────────────────────────────────────────────
    T.append(('header', 'F.', 'Substance Abuse', None))
    T.append(('data', 860))  # Alcohol use disorders
    T.append(('data', 870))  # Drug use disorders
    T.append(('blank',))

    # ── Section G: Others ─────────────────────────────────────────────────────
    T.append(('header', 'G.', 'Others', None))
    T.append(('data', 40))   # STDs excluding HIV
    T.append(('data', 100))  # HIV/AIDS
    T.append(('data', 365))  # Leprosy
    T.append(('data', 810))  # Endocrine, blood, immune disorders
    T.append(('data', 940))  # Neurological conditions
    T.append(('data', 1020)) # Sense organ diseases
    T.append(('data', 1210)) # Digestive diseases
    T.append(('data', 1260)) # Genitourinary diseases
    T.append(('data', 1330)) # Skin diseases
    T.append(('data', 1340)) # Musculoskeletal diseases
    T.append(('data', 1470)) # Oral conditions
    return T


# ── DATA LOADING ──────────────────────────────────────────────────────────────

def load_global_sheet(raw_file, sheet_name):
    """
    Load one 'Global YYYY' sheet from the raw WHO file.

    Returns:
        header_rows  – list of 9 lists (raw rows 0–8, all columns)
        data_lookup  – dict  { ghe_code (int): [float × N_COLS] }
                       Values are the 19 sex × age-group DALYs for that cause.
        pop_vals     – list of N_COLS floats (global population in thousands, raw row 8)
        year_str     – year as a string, e.g. "2021"

    Raw sheet structure:
        Row 1  – title (column A)
        Row 3  – "Region:" label + "Global" (column F)
        Row 4  – "Year:"   label + year value (column F)
        Row 6  – sex header:       Both sexes | Male | Female | Male … | Female …
        Row 7  – age group header: Total | Total | Total | 0-28 days | 1-59 months …
        Row 8  – population row (in thousands)
        Row 9  – column headers: Code | Cause of death | …
        Row 10+– cause data rows
    """
    df = pd.read_excel(raw_file, sheet_name=sheet_name, header=None)
    total_data_rows = max(0, len(df) - 9)
    stride = get_progress_stride(total_data_rows)
    print(f"    Loaded sheet '{sheet_name}' ({len(df):,} total rows, "
          f"{total_data_rows:,} data rows)")

    header_rows = [list(df.iloc[ri]) for ri in range(9)]

    # Year string from row 4 (0-indexed row 3), column F (0-based index 5)
    year_raw = df.iloc[3].iloc[5]
    year_str = str(int(year_raw)) if pd.notna(year_raw) else sheet_name.split()[-1]

    # Population (row 8, 0-indexed row 7) – already expressed in thousands
    pop_row  = df.iloc[7]
    pop_vals = [
        float(pop_row.iloc[DATA_COL_START + ci]) if pd.notna(pop_row.iloc[DATA_COL_START + ci])
        else 0.0
        for ci in range(N_COLS)
    ]

    # Cause data (rows 10+, 0-indexed 9+)
    data_lookup = {}
    for data_idx, ri in enumerate(range(9, len(df)), start=1):
        row     = df.iloc[ri]
        ghe_raw = row.iloc[0]
        if pd.isna(ghe_raw):
            continue
        try:
            ghe_code = int(float(ghe_raw))
        except (ValueError, TypeError):
            continue
        vals = [
            float(row.iloc[DATA_COL_START + ci]) if pd.notna(row.iloc[DATA_COL_START + ci])
            else 0.0
            for ci in range(N_COLS)
        ]
        data_lookup[ghe_code] = vals

        if data_idx == 1 or data_idx == total_data_rows or data_idx % stride == 0:
            print_progress(f"  Parsing {sheet_name}", data_idx, total_data_rows)

    return header_rows, data_lookup, pop_vals, year_str


# ── BUILD OUTPUT ROWS ─────────────────────────────────────────────────────────

def build_global_output_rows(template, data_lookup, pop_vals):
    """
    Convert a template + data_lookup into a list of output rows.

    Each row is a list of (FIXED_COLS + N_COLS) values:
      [ghe_code, col2, col3, col4, col5, spacer,
       val_BothTotal, val_MaleTotal, val_FemTotal,
       val_Male_0-28d, …, val_Female_70+]

    Column layout (0-based):
      0 – GHE code  (int or None)
      1 – col2      (hierarchical letter, e.g. 'I.' – usually None)
      2 – col3      (section letter, e.g. 'A.')
      3 – col4      (section / category name)
      4 – col5      (leaf cause label)
      5 – spacer    (always None)
      6…24 – 19 sex × age-group float values
    """
    output_rows = []

    def get_vals(ghe_code):
        return list(data_lookup.get(ghe_code, [0.0] * N_COLS))

    def calc_vals(calc_fn):
        results = []
        for ci in range(N_COLS):
            per_code = {code: vals[ci] for code, vals in data_lookup.items()}
            results.append(calc_fn(per_code))
        return results

    for entry in template:
        kind = entry[0]

        if kind == 'blank':
            output_rows.append([None] * (FIXED_COLS + N_COLS))

        elif kind == 'pop':
            # Population row – values are in thousands, matching the raw file
            row = [None, None, None, "Population ('000)", None, None] + list(pop_vals)
            output_rows.append(row)

        elif kind == 'header':
            _, col3, col4, col5 = entry
            row = [None, None, col3, col4, col5, None] + [None] * N_COLS
            output_rows.append(row)

        elif kind == 'data':
            _, ghe_code = entry
            col2, col3, col4, col5 = GHE_LABEL_MAP.get(
                ghe_code, (None, None, None, str(ghe_code))
            )
            output_rows.append([ghe_code, col2, col3, col4, col5, None] + get_vals(ghe_code))

        elif kind == 'calc':
            _, label, calc_fn = entry
            output_rows.append([None, None, None, None, label, None] + calc_vals(calc_fn))

    return output_rows


# ── ROW STYLING ───────────────────────────────────────────────────────────────

def _classify_row(row_data):
    """
    Return a style_key string for one output row based on its metadata.

    Row layout assumed:
      [0] ghe_code  [1] col2  [2] col3  [3] col4  [4] col5  [5] spacer  [6+] values

    Row type priority (matches categorise_daly.py):
      blank      – all entries are None
      population – col4 starts with "Population"
      calc       – ghe_code is None and col5 is a non-empty string
      section    – ghe_code is None and col3 or col4 or col5 has text
      summary    – ghe_code in {0, 10, 20, 600, 1510}
      body       – everything else
    """
    if all(v is None for v in row_data):
        return 'blank'

    ghe_code = row_data[0]
    col4     = row_data[3]
    col5     = row_data[4]
    col3     = row_data[2]

    is_population = isinstance(col4, str) and col4.startswith("Population")
    is_calc    = (ghe_code is None and isinstance(col5, str) and not is_population)
    is_section = (ghe_code is None and
                  any(isinstance(v, str) and str(v).strip()
                      for v in (col3, col4, col5)))
    is_summary = ghe_code in {0, 10, 20, 600, 1510}

    if   is_population: return 'population'
    elif is_calc:       return 'calc'
    elif is_section:    return 'section'
    elif is_summary:    return 'summary'
    return 'body'


def style_data_row_global(ws, row_idx, row_data, styles, n_val_cols=N_COLS, is_banded=False):
    """
    Apply formatting to one data row in a global categorised sheet.
    Mirrors style_data_row() from categorise_daly.py with the global column layout.
    """
    row_cells  = list(ws[row_idx])
    meta_cells = row_cells[:FIXED_COLS]
    val_cells  = row_cells[FIXED_COLS:FIXED_COLS + n_val_cols]

    style_key = _classify_row(row_data[:FIXED_COLS + n_val_cols])

    if style_key == 'blank':
        for cell in meta_cells + val_cells:
            apply_cell_style(cell, styles['blank'])
        ws.row_dimensions[row_idx].height = 8
        return

    for cell in meta_cells:
        apply_cell_style(cell, styles[style_key])

    for cell in val_cells:
        apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
        if style_key in ('population', 'calc', 'section', 'summary'):
            cell.fill = styles[style_key]['fill']
        elif is_banded:
            cell.fill = styles['band_fill']

    if style_key == 'body' and is_banded:
        for cell in meta_cells:
            cell.fill = styles['band_fill']

    if   style_key == 'section':              ws.row_dimensions[row_idx].height = 22
    elif style_key in ('population','summary'): ws.row_dimensions[row_idx].height = 20
    else:                                       ws.row_dimensions[row_idx].height = 18


# ── WRITE CATEGORISED SHEET ───────────────────────────────────────────────────

def write_global_categorised_sheet(ws, output_rows, year_str, sheet_label=None):
    """
    Write one 'YYYY Categorised' sheet.

    Layout mirrors the by-country categorised sheets:
      Rows 1–4  : WHO metadata / title
      Row  5    : (blank)
      Row  6    : year label
      Row  7    : main column headers  (fixed col names + sex names per value col)
      Row  8    : sub-headers          (age-group names per value col)
      Row  9+   : data rows

    Value columns (G onwards):
      Col G  = Both sexes, Total
      Col H  = Male, Total
      Col I  = Female, Total
      Cols J–Q = Male by age group (0-28d … 70+)
      Cols R–Y = Female by age group (0-28d … 70+)
    """
    styles = build_style_palette()

    # ── Metadata rows 1–6 ─────────────────────────────────────────────────────
    meta_texts = {
        1: "World Health Organization",
        2: "Department of Data and Analytics",
        3: "Global Health Estimates 2021",
        5: "Global Estimated DALY ('000) by cause, age and sex",
        6: year_str,
    }
    for row_num, text in meta_texts.items():
        cell = ws.cell(row=row_num, column=4, value=text)
        apply_cell_style(cell, styles['title'])
    for ri in range(1, 7):
        ws.row_dimensions[ri].height = 20
    for merge_range in ('D1:F1', 'D2:F2', 'D3:F3', 'D5:F5', 'D6:F6'):
        ws.merge_cells(merge_range)

    # ── Row 7: main column headers ────────────────────────────────────────────
    fixed_headers = ["GHE Code", "", "Cat.", "Section / Category", "Cause", ""]
    for ci, h in enumerate(fixed_headers, start=1):
        apply_cell_style(ws.cell(row=7, column=ci, value=h), styles['header'])
    for ci, (sex, _age) in enumerate(GLOBAL_COLUMNS):
        apply_cell_style(
            ws.cell(row=7, column=FIXED_COLS + 1 + ci, value=sex),
            styles['header']
        )
    ws.row_dimensions[7].height = 34

    # ── Row 8: age-group sub-headers ─────────────────────────────────────────
    for ci in range(FIXED_COLS):
        apply_cell_style(ws.cell(row=8, column=ci + 1, value=""), styles['subheader'])
    for ci, (_sex, age) in enumerate(GLOBAL_COLUMNS):
        apply_cell_style(
            ws.cell(row=8, column=FIXED_COLS + 1 + ci, value=age),
            styles['subheader']
        )
    ws.row_dimensions[8].height = 24

    # ── Data rows from row 9 ─────────────────────────────────────────────────
    current_row   = 9
    stride        = get_progress_stride(len(output_rows))

    for local_idx, dr in enumerate(output_rows, start=1):
        for ci, val in enumerate(dr):
            ws.cell(row=current_row, column=ci + 1, value=val)
        style_data_row_global(ws, current_row, dr, styles,
                               is_banded=(current_row % 2 == 0))
        if sheet_label and (
            local_idx == 1 or local_idx == len(output_rows) or local_idx % stride == 0
        ):
            print_progress(f"  Writing {sheet_label}", local_idx, len(output_rows))
        current_row += 1

    # ── Freeze panes, gridlines, column widths ────────────────────────────────
    ws.freeze_panes = f'{get_column_letter(FIXED_COLS + 1)}9'
    ws.sheet_view.showGridLines = False

    col_widths_fixed = [12, 6, 6, 32, 44, 4]
    for ci, w in enumerate(col_widths_fixed, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ci in range(N_COLS):
        ws.column_dimensions[get_column_letter(FIXED_COLS + 1 + ci)].width = 16


# ── COPY RAW SHEET (for ORIGINAL reference tab) ───────────────────────────────

def copy_and_format_raw_sheet(raw_file, sheet_name, ws_out):
    """
    Copy a raw 'Global YYYY' sheet verbatim into ws_out and apply
    light formatting to distinguish the header rows visually.
    """
    styles = build_style_palette()

    wb_raw    = load_workbook(raw_file, read_only=True, data_only=True)
    ws_raw    = wb_raw[sheet_name]
    total_raw = ws_raw.max_row or 0
    stride    = get_progress_stride(total_raw)

    for row_idx, row in enumerate(ws_raw.iter_rows(values_only=True), start=1):
        ws_out.append(list(row))
        if row_idx == 1 or row_idx == total_raw or row_idx % stride == 0:
            print_progress(f"  Copying {sheet_name}", row_idx, total_raw)
    wb_raw.close()

    max_col = ws_out.max_column

    # Row 1: title
    for cell in ws_out[1][:max_col]:
        if cell.value not in (None, ''):
            apply_cell_style(cell, styles['title'])
    ws_out.row_dimensions[1].height = 22

    # Rows 6–7: sex and age-group headers
    for row_num in (6, 7):
        if ws_out.max_row >= row_num:
            for cell in ws_out[row_num][:max_col]:
                apply_cell_style(cell, styles['header'])
            ws_out.row_dimensions[row_num].height = 28

    # Row 8: population (yellow)
    if ws_out.max_row >= 8:
        for cell in ws_out[8][:max_col]:
            apply_cell_style(cell, styles['population'])
        ws_out.row_dimensions[8].height = 22

    # Row 9: column-label header
    if ws_out.max_row >= 9:
        for cell in ws_out[9][:max_col]:
            apply_cell_style(cell, styles['subheader'])
        ws_out.row_dimensions[9].height = 22

    # Number format for value cells
    for row_num in range(10, ws_out.max_row + 1):
        for col_idx in range(DATA_COL_START + 1, max_col + 1):
            cell = ws_out.cell(row=row_num, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMAT

    # Column widths
    for ci in range(1, DATA_COL_START + 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = 28
    for ci in range(DATA_COL_START + 1, max_col + 1):
        ws_out.column_dimensions[get_column_letter(ci)].width = 18

    ws_out.freeze_panes = f'{get_column_letter(DATA_COL_START + 1)}10'
    ws_out.sheet_view.showGridLines = False


# ── TRENDS SHEET ──────────────────────────────────────────────────────────────

def write_trends_sheet(ws, template, all_year_data, year_labels):
    """
    Write the 'Trends – Both Sexes Total' sheet.

    Rows   : same GHE cause selection as _global_categorised_template()
    Columns: one per year  (Both sexes total only = GLOBAL_COLUMNS index 0)

    This makes it easy to see how each disease burden has changed over 2000–2021.
    """
    styles  = build_style_palette()
    n_years = len(year_labels)

    # ── Title rows ────────────────────────────────────────────────────────────
    ws.cell(row=1, column=4,
            value="WHO Global DALY Trends – Both Sexes, Total (all ages) ('000)")
    apply_cell_style(ws.cell(row=1, column=4), styles['title'])
    ws.merge_cells('D1:F1')
    ws.row_dimensions[1].height = 22

    ws.cell(row=2, column=4,
            value="Selected causes, all six GHE 2021 estimate years")
    apply_cell_style(ws.cell(row=2, column=4), styles['subheader'])
    ws.merge_cells('D2:F2')
    ws.row_dimensions[2].height = 18

    # ── Column headers (row 4) ────────────────────────────────────────────────
    fixed_headers = ["GHE Code", "", "Cat.", "Section / Category", "Cause", ""]
    for ci, h in enumerate(fixed_headers, start=1):
        apply_cell_style(ws.cell(row=4, column=ci, value=h), styles['header'])
    for yi, yr in enumerate(year_labels):
        apply_cell_style(ws.cell(row=4, column=FIXED_COLS + 1 + yi, value=yr), styles['header'])
    ws.row_dimensions[4].height = 28

    # ── Sub-header (row 5) ────────────────────────────────────────────────────
    for ci in range(FIXED_COLS):
        apply_cell_style(ws.cell(row=5, column=ci + 1, value=""), styles['subheader'])
    for yi in range(n_years):
        apply_cell_style(
            ws.cell(row=5, column=FIXED_COLS + 1 + yi, value="Both sexes\nTotal"),
            styles['subheader']
        )
    ws.row_dimensions[5].height = 28

    # ── Data rows from row 6 ──────────────────────────────────────────────────
    current_row = 6
    for entry in template:
        kind = entry[0]

        if kind == 'pop':
            continue  # population row omitted in trends view

        if kind == 'blank':
            for ci in range(FIXED_COLS + n_years):
                apply_cell_style(ws.cell(row=current_row, column=ci + 1), styles['blank'])
            ws.row_dimensions[current_row].height = 8
            current_row += 1
            continue

        # Build metadata and year values based on kind
        if kind == 'header':
            _, col3, col4, col5 = entry
            base_meta  = [None, None, col3, col4, col5, None]
            year_vals  = [None] * n_years

        elif kind == 'data':
            _, ghe_code = entry
            col2, col3, col4, col5 = GHE_LABEL_MAP.get(
                ghe_code, (None, None, None, str(ghe_code))
            )
            base_meta = [ghe_code, col2, col3, col4, col5, None]
            year_vals = [
                all_year_data[yr].get(ghe_code, [0.0] * N_COLS)[0]
                for yr in year_labels
            ]

        elif kind == 'calc':
            _, label, calc_fn = entry
            base_meta = [None, None, None, None, label, None]
            year_vals = []
            for yr in year_labels:
                per_code = {code: vals[0] for code, vals in all_year_data[yr].items()}
                year_vals.append(calc_fn(per_code))
        else:
            continue

        full_row = base_meta + year_vals
        for ci, val in enumerate(full_row):
            ws.cell(row=current_row, column=ci + 1, value=val)

        # Style this row using the same classification logic
        style_key = _classify_row(base_meta + [None] * N_COLS)
        for ci in range(FIXED_COLS):
            apply_cell_style(ws.cell(row=current_row, column=ci + 1), styles[style_key])
        for yi in range(n_years):
            cell = ws.cell(row=current_row, column=FIXED_COLS + 1 + yi)
            apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
            if style_key in ('calc', 'section', 'summary') and year_vals[yi] is not None:
                cell.fill = styles[style_key]['fill']

        row_height = 22 if style_key == 'section' else (20 if style_key == 'summary' else 18)
        ws.row_dimensions[current_row].height = row_height
        current_row += 1

    # ── Freeze, widths, gridlines ─────────────────────────────────────────────
    ws.freeze_panes = f'{get_column_letter(FIXED_COLS + 1)}6'
    ws.sheet_view.showGridLines = False

    col_widths_fixed = [12, 6, 6, 32, 44, 4]
    for ci, w in enumerate(col_widths_fixed, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for yi in range(n_years):
        ws.column_dimensions[get_column_letter(FIXED_COLS + 1 + yi)].width = 16


# ── MPNc GLOBAL SHEET ─────────────────────────────────────────────────────────

MPNC_GHE_LABELS = {
    420:  "Maternal conditions",
    490:  "Neonatal conditions  [computed: 500+510+520+530]",
    500:  "Preterm birth complications",
    510:  "Birth asphyxia and birth trauma",
    520:  "Neonatal sepsis and infections",
    530:  "Other neonatal conditions",
    540:  "Nutritional deficiencies  [computed: 550+…+590]",
    550:  "Protein-energy malnutrition",
    560:  "Iodine deficiency",
    570:  "Vitamin A deficiency",
    580:  "Iron-deficiency anaemia",
    590:  "Other nutritional deficiencies",
    1505: "Sudden infant death syndrome",
}

MPNC_DISPLAY_ORDER = [420, 490, 500, 510, 520, 530, 540, 550, 560, 570, 580, 590, 1505]


def _mpnc_get_col0(ghe_code, data_lookup):
    """
    Return the Both-sexes total (column index 0) for one MPNc GHE code.
    GHE 490 and 540 are aggregate codes not present in the raw file and must
    be computed from their sub-codes.
    """
    if ghe_code == 490:
        return sum(data_lookup.get(sc, [0.0])[0] for sc in [500, 510, 520, 530])
    if ghe_code == 540:
        return sum(data_lookup.get(sc, [0.0])[0] for sc in [550, 560, 570, 580, 590])
    return data_lookup.get(ghe_code, [0.0])[0]


def _mpnc_total_col0(data_lookup):
    """Total MPNc = sum of MPNc leaf codes (avoiding double-counting 490/540)."""
    leaf = [420, 500, 510, 520, 530, 550, 560, 570, 580, 590, 1505]
    return sum(data_lookup.get(c, [0.0])[0] for c in leaf)


def write_mpnc_global_sheet(ws, all_year_data, year_labels):
    """
    Write the 'MPNc – Global' sheet.

    Rows   : Total MPNc (summary) + each individual MPNc GHE code
    Columns: one per year  (Both sexes total only)

    Mirrors the MPNc and MPNc (2) tabs from the by-country output but adapted
    for the global, multi-year format.
    """
    styles  = build_style_palette()
    n_years = len(year_labels)

    # ── Title rows ────────────────────────────────────────────────────────────
    ws.cell(row=1, column=4,
            value="WHO Global DALY – Maternal, Perinatal and Neonatal Conditions (MPNc)")
    apply_cell_style(ws.cell(row=1, column=4), styles['title'])
    ws.merge_cells('D1:F1')
    ws.row_dimensions[1].height = 22

    ws.cell(row=2, column=4,
            value="Both sexes, Total (all ages) — DALYs ('000).  "
                  "GHE 490 & 540 computed from sub-codes.")
    apply_cell_style(ws.cell(row=2, column=4), styles['subheader'])
    ws.merge_cells('D2:F2')
    ws.row_dimensions[2].height = 18

    # ── Column headers (row 4) ────────────────────────────────────────────────
    fixed_headers = ["GHE Code", "", "", "MPNc Cause", "", ""]
    for ci, h in enumerate(fixed_headers, start=1):
        apply_cell_style(ws.cell(row=4, column=ci, value=h), styles['header'])
    for yi, yr in enumerate(year_labels):
        apply_cell_style(ws.cell(row=4, column=FIXED_COLS + 1 + yi, value=yr), styles['header'])
    ws.row_dimensions[4].height = 28

    # ── Sub-header (row 5) ────────────────────────────────────────────────────
    for ci in range(FIXED_COLS):
        apply_cell_style(ws.cell(row=5, column=ci + 1, value=""), styles['subheader'])
    for yi in range(n_years):
        apply_cell_style(
            ws.cell(row=5, column=FIXED_COLS + 1 + yi, value="Both sexes\nTotal ('000)"),
            styles['subheader']
        )
    ws.row_dimensions[5].height = 28

    current_row = 6

    # ── Total MPNc summary row ────────────────────────────────────────────────
    ws.cell(row=current_row, column=4, value="Total MPNc")
    for ci in range(FIXED_COLS):
        apply_cell_style(ws.cell(row=current_row, column=ci + 1), styles['summary'])
    for yi, yr in enumerate(year_labels):
        cell = ws.cell(row=current_row, column=FIXED_COLS + 1 + yi,
                       value=_mpnc_total_col0(all_year_data[yr]))
        apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
        cell.fill = styles['summary']['fill']
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # ── Blank spacer ─────────────────────────────────────────────────────────
    for ci in range(FIXED_COLS + n_years):
        apply_cell_style(ws.cell(row=current_row, column=ci + 1), styles['blank'])
    ws.row_dimensions[current_row].height = 8
    current_row += 1

    # ── Detail rows ───────────────────────────────────────────────────────────
    for ghe_code in MPNC_DISPLAY_ORDER:
        label      = MPNC_GHE_LABELS.get(ghe_code, str(ghe_code))
        is_computed = ghe_code in (490, 540)
        style_key  = 'calc' if is_computed else 'body'

        ws.cell(row=current_row, column=1,
                value=None if is_computed else ghe_code)
        ws.cell(row=current_row, column=4, value=label)
        for ci in range(FIXED_COLS):
            apply_cell_style(ws.cell(row=current_row, column=ci + 1), styles[style_key])
        for yi, yr in enumerate(year_labels):
            cell = ws.cell(row=current_row, column=FIXED_COLS + 1 + yi,
                           value=_mpnc_get_col0(ghe_code, all_year_data[yr]))
            apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
            if is_computed:
                cell.fill = styles['calc']['fill']
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # ── Freeze, widths, gridlines ─────────────────────────────────────────────
    ws.freeze_panes = f'{get_column_letter(FIXED_COLS + 1)}6'
    ws.sheet_view.showGridLines = False

    col_widths_fixed = [12, 6, 6, 48, 6, 4]
    for ci, w in enumerate(col_widths_fixed, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for yi in range(n_years):
        ws.column_dimensions[get_column_letter(FIXED_COLS + 1 + yi)].width = 16


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print(f"Loading raw file: {RAW_FILE}")
    if not os.path.exists(RAW_FILE):
        raise FileNotFoundError(f"Raw file not found: {RAW_FILE}")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)   # remove the default empty sheet

    template            = _global_categorised_template()
    all_year_data       = {}   # year_str  ->  data_lookup
    year_labels_ordered = []   # years in processing order

    total_sheets  = 2 + len(YEAR_SHEETS) + 2   # Top20 + Summary + categorised × 6 + trends + MPNc
    sheet_counter = 0

    # ── Sheet 1: Top20 (copied verbatim from raw file) ────────────────────────
    sheet_counter += 1
    print(f"\n[{sheet_counter}/{total_sheets}] Copying 'Top20' ...")
    ws_top20 = wb_out.create_sheet("Top20")
    copy_and_format_raw_sheet(RAW_FILE, "Top20", ws_top20)
    print(f"  ✓ Done.")

    # ── Sheet 2: Summary (copied verbatim from raw file) ─────────────────────
    sheet_counter += 1
    print(f"\n[{sheet_counter}/{total_sheets}] Copying 'Summary' ...")
    ws_summary = wb_out.create_sheet("Summary")
    copy_and_format_raw_sheet(RAW_FILE, "Summary", ws_summary)
    print(f"  ✓ Done.")

    # ── Sheets 3–8: Categorised sheet per year ────────────────────────────────
    for raw_sheet_name, year_hint in YEAR_SHEETS:
        sheet_counter += 1
        sheet_tab     = f"{year_hint} Categorised"
        print(f"\n[{sheet_counter}/{total_sheets}] Processing {raw_sheet_name} → '{sheet_tab}' ...")
        header_rows, data_lookup, pop_vals, year_str = load_global_sheet(
            RAW_FILE, raw_sheet_name
        )
        all_year_data[year_str]    = data_lookup
        year_labels_ordered.append(year_str)

        ws_cat      = wb_out.create_sheet(sheet_tab)
        output_rows = build_global_output_rows(template, data_lookup, pop_vals)
        write_global_categorised_sheet(ws_cat, output_rows, year_str, sheet_label=sheet_tab)
        print(f"  ✓ Done. ({len(output_rows)} data rows, {len(data_lookup)} GHE codes loaded)")

    # ── Sheet 8: Trends – Both Sexes Total ───────────────────────────────────
    sheet_counter += 1
    print(f"\n[{sheet_counter}/{total_sheets}] Writing 'Trends – Both Sexes Total' ...")
    ws_trends = wb_out.create_sheet("Trends – Both Sexes Total")
    write_trends_sheet(ws_trends, template, all_year_data, year_labels_ordered)
    print(f"  ✓ Done.")

    # ── Sheet 9: MPNc – Global ────────────────────────────────────────────────
    sheet_counter += 1
    print(f"\n[{sheet_counter}/{total_sheets}] Writing 'MPNc – Global' ...")
    ws_mpnc = wb_out.create_sheet("MPNc – Global")
    write_mpnc_global_sheet(ws_mpnc, all_year_data, year_labels_ordered)
    print(f"  ✓ Done.")

    # ── Save ──────────────────────────────────────────────────────────────────
    print(f"\nSaving workbook to {OUTPUT_FILE} ...")
    wb_out.save(OUTPUT_FILE)

    print(f"\n{'─'*60}")
    print(f"Done!  Output: {OUTPUT_FILE}")
    print(f"Sheets written ({len(wb_out.sheetnames)}):")
    for sn in wb_out.sheetnames:
        print(f"  – {sn}")
    print(f"{'─'*60}")


if __name__ == "__main__":
    main()
