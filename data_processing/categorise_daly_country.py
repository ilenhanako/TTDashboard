"""
WHO DALY Categorisation Script
================================
Transforms raw WHO GHE DALY-by-country Excel files into the standardised
categorised format used for analysis (matching ghe2021daly_categorised_finalised.xlsx).

Usage:
    python categorise_daly.py

Configure RAW_FILE and OUTPUT_FILE below. No command-line arguments needed.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
import copy, re, os


NUMBER_FORMAT = '#,##0.0'
PROGRESS_CHECKPOINTS = 8
MIN_PROGRESS_STRIDE = 25


def get_progress_stride(total_rows, checkpoints=PROGRESS_CHECKPOINTS, min_stride=MIN_PROGRESS_STRIDE):
    """Return a sensible progress-report interval for row-based loops."""
    if total_rows <= 0:
        return 1
    return max(1, min(total_rows, max(min_stride, total_rows // checkpoints)))


def print_progress(label, current, total):
    """Emit a consistent terminal progress message for long-running steps."""
    if total <= 0:
        print(f"    {label}: {current:,}")
        return

    percentage = (current / total) * 100
    print(f"    {label}: {current:,}/{total:,} rows ({percentage:5.1f}%)")


_STYLE_CACHE = None


def build_style_palette():
    """Centralised workbook style definitions for readable output sheets (cached)."""
    global _STYLE_CACHE
    if _STYLE_CACHE is not None:
        return _STYLE_CACHE
    thin_gray = Side(style='thin', color='D9E2F3')
    medium_blue = Side(style='medium', color='5B9BD5')
    strong_navy = '1F4E78'

    palette = {
        'title': {
            'font': Font(bold=True, size=12, color=strong_navy),
            'fill': PatternFill('solid', fgColor='EAF3FF'),
            'alignment': Alignment(horizontal='left', vertical='center'),
        },
        'header': {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill('solid', fgColor=strong_navy),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(left=medium_blue, right=medium_blue, top=medium_blue, bottom=medium_blue),
        },
        'subheader': {
            'font': Font(bold=True, color=strong_navy),
            'fill': PatternFill('solid', fgColor='DCE6F1'),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray),
        },
        'section': {
            'font': Font(bold=True, color=strong_navy),
            'fill': PatternFill('solid', fgColor='D9EAF7'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray),
        },
        'summary': {
            'font': Font(bold=True, color='203864'),
            'fill': PatternFill('solid', fgColor='EAF2F8'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray),
        },
        'population': {
            'font': Font(bold=True, color='7F6000'),
            'fill': PatternFill('solid', fgColor='FFF2CC'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray),
        },
        'calc': {
            'font': Font(bold=True, italic=True, color='385723'),
            'fill': PatternFill('solid', fgColor='E2F0D9'),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray),
        },
        'body': {
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray),
        },
        'numeric': {
            'alignment': Alignment(horizontal='right', vertical='center'),
            'border': Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray),
        },
        'blank': {
            'fill': PatternFill('solid', fgColor='F7F9FC'),
        },
        'band_fill': PatternFill('solid', fgColor='F8FBFF'),
    }
    _STYLE_CACHE = palette
    return palette


def apply_cell_style(cell, style_def=None, number_format=None):
    """Apply a partial style dict to a cell. openpyxl styles are immutable so direct assignment is safe."""
    if style_def:
        if 'font' in style_def:
            cell.font = style_def['font']
        if 'fill' in style_def:
            cell.fill = style_def['fill']
        if 'alignment' in style_def:
            cell.alignment = style_def['alignment']
        if 'border' in style_def:
            cell.border = style_def['border']
    if number_format is not None:
        cell.number_format = number_format


def style_data_row(ws, row_idx, row_data, country_count, styles, is_banded=False):
    """Apply row-level formatting based on the semantic row type."""
    row_cells = list(ws[row_idx])
    meta_cells = row_cells[:7]
    value_cells = row_cells[7:7 + country_count]

    if all(val is None for val in row_data[:7]) and all(val is None for val in row_data[7:]):
        for cell in row_cells[:7 + country_count]:
            apply_cell_style(cell, styles['blank'])
        ws.row_dimensions[row_idx].height = 8
        return

    label_c3 = row_data[3]
    label_c4 = row_data[4]
    label_c5 = row_data[5]
    is_population = isinstance(label_c3, str) and label_c3.startswith("Population")
    is_calc = row_data[1] is None and isinstance(label_c5, str) and not is_population
    is_section = row_data[0] is None and any(isinstance(val, str) and val.strip() for val in (label_c3, label_c4, label_c5))
    is_summary = row_data[1] in {0, 10, 20, 600, 1510}

    if is_population:
        style_key = 'population'
    elif is_calc:
        style_key = 'calc'
    elif is_section:
        style_key = 'section'
    elif is_summary:
        style_key = 'summary'
    else:
        style_key = 'body'

    for cell in meta_cells:
        apply_cell_style(cell, styles[style_key])

    for cell in value_cells:
        apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
        if style_key in ('population', 'calc', 'section', 'summary') and 'fill' in styles[style_key]:
            cell.fill = styles[style_key]['fill']
        elif is_banded:
            cell.fill = styles['band_fill']

    if style_key == 'body' and is_banded:
        for cell in meta_cells:
            cell.fill = styles['band_fill']

    if is_section:
        ws.row_dimensions[row_idx].height = 22
    elif is_population or is_summary:
        ws.row_dimensions[row_idx].height = 20
    else:
        ws.row_dimensions[row_idx].height = 18


def format_full_sheet(ws, progress_label=None):
    """Apply the workbook styling to a copied raw sheet without changing its data."""
    styles = build_style_palette()
    max_col = ws.max_column
    country_count = max(0, max_col - 7)

    for row_idx in range(1, min(ws.max_row, 6) + 1):
        for cell in ws[row_idx][3:6]:
            if cell.value not in (None, ''):
                apply_cell_style(cell, styles['title'])

    if ws.max_row >= 7:
        for cell in ws[7][:max_col]:
            apply_cell_style(cell, styles['header'])
        ws.row_dimensions[7].height = 34

    for row_idx in (8, 9):
        if ws.max_row >= row_idx:
            for cell in ws[row_idx][:max_col]:
                apply_cell_style(cell, styles['subheader'])
            ws.row_dimensions[row_idx].height = 24

    data_row_count = max(0, ws.max_row - 9)
    progress_stride = get_progress_stride(data_row_count)

    for row_idx in range(10, ws.max_row + 1):
        row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
        is_banded = row_idx % 2 == 0
        style_data_row(ws, row_idx, row_data, country_count, styles, is_banded=is_banded)

        processed_rows = row_idx - 9
        if progress_label and (
            processed_rows == 1 or
            processed_rows == data_row_count or
            processed_rows % progress_stride == 0
        ):
            print_progress(progress_label, processed_rows, data_row_count)

        for col_idx in range(8, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMAT

    ws.freeze_panes = 'H10'
    ws.sheet_view.showGridLines = False

    for merge_range in ('D1:F1', 'D2:F2', 'D3:F3', 'D5:F5', 'D6:F6'):
        ws.merge_cells(merge_range)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 38
    ws.column_dimensions['F'].width = 42
    ws.column_dimensions['G'].width = 20

    for col_idx in range(8, max_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 14

    for row_idx in range(1, min(ws.max_row, 6) + 1):
        ws.row_dimensions[row_idx].height = 20

# CONFIGURATION
RAW_FILE   = "GHE_DALYOriginal/country/ghe2021_daly_bycountry_2000.xlsx"
OUTPUT_FILE = "GHE_DALYFinal/country/ghe2000_daly_categorised_formatted.xlsx"

YEAR = "2000"

TARGET_COUNTRIES = [
    "Bangladesh", "Brunei Darussalam", "Cambodia", "China", "India",
    "Indonesia", "Japan", "Lao People's Democratic Republic", "Malaysia",
    "Myanmar", "Philippines", "Republic of Korea", "Singapore", "Viet Nam",
]

MPNC_FOCUSED_COUNTRIES = [
    "Cambodia", "China", "India", "Indonesia",
    "Lao People's Democratic Republic", "Philippines", "Viet Nam",
]

MPNC_AGE_GROUPS = [
    ("All ",                     "All ages"),
    ("Toddler (0-4)",            "0-4"),
    ("School Age Child (5-14)",  "5-14"),
    ("Adolescent/YA (15-29)",    "15-29"),
]

MPNC_GHE_LABELS = {
    420:  "Maternal conditions",
    490:  "Neonatal conditions",
    500:  "Preterm birth complications",
    510:  "Birth asphyxia and birth trauma",
    520:  "Neonatal sepsis and infections",
    530:  "Other neonatal conditions",
    540:  "Nutritional deficiencies",
    550:  "Protein-energy malnutrition",
    560:  "Iodine deficiency",
    570:  "Vitamin A deficiency",
    580:  "Iron-deficiency anaemia",
    590:  "Other nutritional deficiencies",
    1505: "Sudden infant death syndrome",
}

MPNC_DISPLAY_ORDER = [420, 490, 500, 510, 520, 530, 540, 550, 560, 570, 580, 590, 1505]

MPNC_FOCUSED_ISO = {
    "Cambodia":                          "KHM",
    "China":                             "CHN",
    "India":                             "IND",
    "Indonesia":                         "IDN",
    "Lao People's Democratic Republic":  "LAO",
    "Philippines":                       "PHL",
    "Viet Nam":                          "VNM",
}

COUNTRY_INCOME_GROUPS = {
    "Bangladesh":                         4,
    "Brunei Darussalam":                  1,
    "Cambodia":                           4,
    "China":                              4,
    "India":                              4,
    "Indonesia":                          4,
    "Japan":                              1,
    "Lao People's Democratic Republic":   4,
    "Malaysia":                           3,
    "Myanmar":                            4,
    "Philippines":                        2,
    "Republic of Korea":                  1,
    "Singapore":                          1,
    "Viet Nam":                           4,
}


# ROW TEMPLATES

def _parasitic_excl(d):
    """Parasitic & vector diseases excl. Malaria (220) and Dengue (300)."""
    return d.get(210, 0) - d.get(220, 0) - d.get(300, 0)

def _cv_excl_stroke(d):
    """Cardiovascular diseases excl. Stroke = GHE 1100 – GHE 1140."""
    return d.get(1100, 0) - d.get(1140, 0)

def _neoplasms(d):
    """Neoplasms = Malignant neoplasms (610) + Other neoplasms (790)."""
    return d.get(610, 0) + d.get(790, 0)


def _persons_all_ages_template():
    T = []
    T.append(('data', 'Persons', 0))
    T.append(('blank',))
    T.append(('data', 'Persons', 10))
    T.append(('data', 'Persons', 20))
    T.append(('data', 'Persons', 30))
    T.append(('data', 'Persons', 110))
    T.append(('data', 'Persons', 185))
    T.append(('data', 'Persons', 220))
    T.append(('data', 'Persons', 300))
    T.append(('calc', 'Persons',
              'Parasitic and vector diseases excl Malaria and Dengue',
              None, None, _parasitic_excl))
    T.append(('data', 'Persons', 370))
    T.append(('data', 'Persons', 380))
    T.append(('data', 'Persons', 1700))
    T.append(('data', 'Persons', 170))
    T.append(('data', 'Persons', 180))
    T.append(('data', 'Persons', 120))
    T.append(('data', 'Persons', 330))
    T.append(('blank',))
    T.append(('header', 'B.', 'Maternal, Perinatal and Nutritional Deficiencies Conditions', None))
    T.append(('data', 'Persons', 420))
    T.append(('data', 'Persons', 500))
    T.append(('data', 'Persons', 510))
    T.append(('data', 'Persons', 520))
    T.append(('data', 'Persons', 530))
    T.append(('data', 'Persons', 550))
    T.append(('data', 'Persons', 560))
    T.append(('data', 'Persons', 570))
    T.append(('data', 'Persons', 580))
    T.append(('data', 'Persons', 590))
    T.append(('data', 'Persons', 1505))
    T.append(('blank',))
    T.append(('data', 'Persons', 600))
    T.append(('data', 'Persons', 800))
    T.append(('calc', 'Persons',
              'Cardiovascular diseases excl Stroke',
              None, None, _cv_excl_stroke))
    T.append(('data', 'Persons', 1140))
    T.append(('data', 'Persons', 1180))
    T.append(('data', 'Persons', 1190))
    T.append(('data', 'Persons', 1200))
    T.append(('calc', 'Persons',
              'Neoplasms', None, None, _neoplasms))
    T.append(('blank',))
    T.append(('data', 'Persons', 1510))
    T.append(('data', 'Persons', 1520))
    T.append(('data', 'Persons', 1630))
    T.append(('data', 'Persons', 1620))
    T.append(('blank',))
    T.append(('header', 'E.', 'Mental Disorders & Wellbeing', None))
    T.append(('data', 'Persons', 1610))
    T.append(('data', 'Persons', 830))
    T.append(('data', 'Persons', 880))
    T.append(('data', 'Persons', 890))
    T.append(('data', 'Persons', 911))
    T.append(('data', 'Persons', 912))
    T.append(('blank',))
    T.append(('header', 'F.', 'Substance Abuse', None))
    T.append(('data', 'Persons', 860))
    T.append(('data', 'Persons', 870))
    T.append(('blank',))
    T.append(('header', 'G.', 'Others', None))
    T.append(('data', 'Persons', 40))
    T.append(('data', 'Persons', 100))
    T.append(('data', 'Persons', 365))
    T.append(('data', 'Persons', 810))
    T.append(('data', 'Persons', 940))
    T.append(('data', 'Persons', 1020))
    T.append(('data', 'Persons', 1210))
    T.append(('data', 'Persons', 1260))
    T.append(('data', 'Persons', 1330))
    T.append(('data', 'Persons', 1340))
    T.append(('data', 'Persons', 1470))
    T.append(('data', 'Persons', 940))
    return T


def _sex_section_template(sex):
    """Single-sex section used inside age-group sheets and the Gender sheet."""
    T = []
    T.append(('data', sex, 0))
    T.append(('blank',))
    T.append(('data', sex, 10))
    T.append(('data', sex, 20))
    T.append(('data', sex, 30))
    T.append(('data', sex, 110))
    T.append(('data', sex, 120))
    T.append(('data', sex, 170))
    T.append(('data', sex, 180))
    T.append(('data', sex, 185))
    T.append(('data', sex, 220))
    T.append(('data', sex, 300))
    T.append(('calc', sex,
              'Parasitic and vector diseases excl Malaria and Dengue',
              None, None, _parasitic_excl))
    T.append(('data', sex, 330))
    T.append(('data', sex, 370))
    T.append(('data', sex, 380))
    T.append(('data', sex, 1700))
    T.append(('blank',))
    T.append(('header', 'B.', 'Maternal, Perinatal and Nutritional Conditions', None))
    T.append(('data', sex, 420))
    T.append(('data', sex, 500))
    T.append(('data', sex, 510))
    T.append(('data', sex, 520))
    T.append(('data', sex, 530))
    T.append(('data', sex, 550))
    T.append(('data', sex, 560))
    T.append(('data', sex, 570))
    T.append(('data', sex, 580))
    T.append(('data', sex, 590))
    T.append(('data', sex, 1505))
    T.append(('blank',))
    T.append(('data', sex, 600))
    T.append(('calc', sex, 'Neoplasms', None, None, _neoplasms))
    T.append(('data', sex, 800))
    T.append(('data', sex, 940))
    T.append(('data', sex, 1020))
    T.append(('data', sex, 1140))
    T.append(('calc', sex, 'Cardiovascular diseases excl Stroke', None, None, _cv_excl_stroke))
    T.append(('data', sex, 1180))
    T.append(('data', sex, 1190))
    T.append(('data', sex, 1200))
    T.append(('blank',))
    T.append(('data', sex, 1510))
    T.append(('data', sex, 1520))
    T.append(('data', sex, 1620))
    T.append(('data', sex, 1630))
    T.append(('blank',))
    T.append(('header', 'E.', 'Mental Disorders and Wellbeing', None))
    T.append(('data', sex, 830))
    T.append(('data', sex, 880))
    T.append(('data', sex, 890))
    T.append(('data', sex, 911))
    T.append(('data', sex, 912))
    T.append(('data', sex, 1610))
    T.append(('blank',))
    T.append(('header', 'F.', 'Substance Abuse', None))
    T.append(('data', sex, 860))
    T.append(('data', sex, 870))
    T.append(('blank',))
    T.append(('header', 'G.', 'Others', None))
    T.append(('data', sex, 40))
    T.append(('data', sex, 100))
    T.append(('data', sex, 365))
    T.append(('data', sex, 810))
    T.append(('data', sex, 1210))
    T.append(('data', sex, 1260))
    T.append(('data', sex, 1330))
    T.append(('data', sex, 1340))
    T.append(('data', sex, 1470))
    return T


def _age_group_template():
    """Template for age-specific sheets: Persons then Males then Females."""
    T = []
    for sex in ('Persons', 'Males', 'Females'):
        T.extend(_sex_section_template(sex))
        T.append(('blank',))
    return T


# DATA LOADING HELPERS

def load_raw_sheet(raw_file, sheet_name):
    """
    Load one sheet from the raw WHO file.
    Returns:
        header_rows  – list of lists (rows 0–8, all columns)
        country_cols – dict  country_name -> column index in raw DataFrame
        data_lookup  – dict  (sex, ghe_code) -> {country_name: value, ...}
        pop_lookup   – dict  sex -> {country_name: population}
    """
    df = pd.read_excel(raw_file, sheet_name=sheet_name, header=None)
    total_data_rows = max(0, len(df) - 9)
    progress_stride = get_progress_stride(total_data_rows)
    print(f"    Loaded sheet '{sheet_name}' into memory with {len(df):,} total rows")

    header_row_idx = 6
    country_row    = df.iloc[header_row_idx]
    iso_row        = df.iloc[7]

    country_cols = {}
    for ci, val in enumerate(country_row):
        if pd.notna(val) and ci > 6:
            country_cols[str(val).strip()] = ci

    data_lookup = {}
    pop_lookup  = {}

    for data_idx, ri in enumerate(range(9, len(df)), start=1):
        row     = df.iloc[ri]
        sex     = row.iloc[0]
        ghe_raw = row.iloc[1]
        col3    = row.iloc[3]

        if data_idx == 1 or data_idx == total_data_rows or data_idx % progress_stride == 0:
            print_progress(f"Parsing {sheet_name}", data_idx, total_data_rows)

        if pd.isna(sex):
            continue
        sex = str(sex).strip()

        if pd.isna(ghe_raw) and pd.notna(col3) and str(col3).startswith('Population'):
            country_vals = {}
            for cn, ci in country_cols.items():
                country_vals[cn] = row.iloc[ci]
            pop_lookup[sex] = country_vals
            continue

        if pd.isna(ghe_raw):
            continue

        try:
            ghe_code = int(float(ghe_raw))
        except (ValueError, TypeError):
            continue

        country_vals = {}
        for cn, ci in country_cols.items():
            v = row.iloc[ci]
            country_vals[cn] = float(v) if pd.notna(v) else 0.0

        data_lookup[(sex, ghe_code)] = country_vals

    header_rows = []
    for ri in range(9):
        header_rows.append(list(df.iloc[ri]))

    return header_rows, country_cols, data_lookup, pop_lookup


def get_label_for_ghe(raw_data_lookup, ghe_code, sex='Persons'):
    """Return the human-readable label stored in the raw data for a GHE code."""
    return None


# LABEL LOOKUP
GHE_LABEL_MAP = {
    0:    (None,  'All Causes',   None, None),
    10:   (None,  'Communicable, maternal, perinatal and nutritional conditions', None, None),
    20:   (None,  'A.',  'Infectious and parasitic diseases', None),
    30:   (None,  None,  None,   'Tuberculosis'),
    40:   (None,  None,  None,   'STDs excluding HIV'),
    100:  (None,  None,  None,   'HIV/AIDS'),
    110:  (None,  None,  None,   'Diarrhoeal diseases'),
    120:  (None,  None,  None,   'Childhood-cluster diseases'),
    170:  (None,  None,  None,   'Meningitis'),
    180:  (None,  None,  None,   'Encephalitis'),
    185:  (None,  None,  None,   'Hepatitis'),
    210:  (None,  None,  None,   'Parasitic and vector diseases'),
    220:  (None,  None,  None,   'Malaria'),
    300:  (None,  None,  None,   'Dengue'),
    330:  (None,  None,  None,   'Intestinal nematode infections'),
    365:  (None,  None,  None,   'Leprosy'),
    370:  (None,  None,  None,   'Other infectious diseases'),
    380:  (None,  None,  None,   'Respiratory Infectious'),
    420:  (None,  None,  None,   'Maternal Conditions'),
    500:  (None,  None,  None,   'Preterm birth complications'),
    510:  (None,  None,  None,   'Birth asphyxia and birth trauma'),
    520:  (None,  None,  None,   'Neonatal sepsis and infections'),
    530:  (None,  None,  None,   'Other neonatal conditions'),
    550:  (None,  None,  None,   'Protein-energy malnutrition'),
    560:  (None,  None,  None,   'Iodine deficiency'),
    570:  (None,  None,  None,   'Vitamin A deficiency'),
    580:  (None,  None,  None,   'Iron-deficiency anaemia'),
    590:  (None,  None,  None,   'Other nutritional deficiencies'),
    600:  (None,  'C.',  'Noncommunicable diseases', None),
    610:  (None,  None,  None,   'Malignant neoplasms'),
    790:  (None,  None,  None,   'Other neoplasms'),
    800:  (None,  None,  None,   'Diabetes mellitus'),
    810:  (None,  None,  None,   'Endocrine, blood, immune disorders'),
    830:  (None,  None,  None,   'Depressive Disorders'),
    860:  (None,  None,  None,   'Alcohol use disorders'),
    870:  (None,  None,  None,   'Drug use disorders'),
    880:  (None,  None,  None,   'Anxiety disorders'),
    890:  (None,  None,  None,   'Eating disorders'),
    911:  (None,  None,  None,   'Attention deficit/hyperactivity syndrome'),
    912:  (None,  None,  None,   'Conduct disorder'),
    940:  (None,  None,  None,   'Neurological conditions'),
    1020: (None,  None,  None,   'Sense organ diseases'),
    1100: (None,  None,  None,   'Cardiovascular diseases'),
    1140: (None,  None,  None,   'Stroke'),
    1180: (None,  None,  None,   'Chronic obstructive pulmonary disease'),
    1190: (None,  None,  None,   'Asthma'),
    1200: (None,  None,  None,   'Other respiratory diseases'),
    1210: (None,  None,  None,   'Digestive diseases'),
    1260: (None,  None,  None,   'Genitourinary diseases'),
    1330: (None,  None,  None,   'Skin diseases'),
    1340: (None,  None,  None,   'Musculoskeletal diseases'),
    1470: (None,  None,  None,   'Oral conditions'),
    1505: (None,  None,  None,   'Sudden infant death syndrome'),
    1510: (None,  'D.',  'Injuries incl War', None),
    1520: (None,  None,  None,   'Unintentional injuries'),
    1610: (None,  None,  None,   'Self-harm'),
    1620: (None,  None,  None,   'Interpersonal violence'),
    1630: (None,  None,  None,   'Collective violence and legal intervention'),
    1700: (None,  None,  None,   'Other COVID-19 pandemic-related outcomes'),
}


# BUILD OUTPUT ROWS

def build_output_rows(template, data_lookup, pop_lookup, target_countries):
    """
    Given a template and the loaded raw data, produce a list of rows.
    Each row is a list:  [Sex, GHE_code, col2, col3, col4, col5,
                          <label_placeholder>, iso_placeholder,
                          val_country1, val_country2, ...]
    (The first 7 entries mirror the raw-file column structure exactly.)
    We keep only 9 header-column slots (0..8) where:
        0 = Sex, 1 = GHE code, 2 = col2, 3 = col3, 4 = col4, 5 = col5,
        6 = Country-or-area label (col header only), 7+ = country values
    """
    n_countries = len(target_countries)
    output_rows = []

    def get_vals(sex, ghe_code):
        return data_lookup.get((sex, ghe_code), {c: 0.0 for c in target_countries})

    def extract_country_vals(sex, ghe_code):
        vals = get_vals(sex, ghe_code)
        return [vals.get(c, 0.0) for c in target_countries]

    def calc_country_vals(sex, calc_fn):
        """Run calc_fn({ghe_code: value_for_country}) for each country."""
        sex_codes = {code: vals for (s, code), vals in data_lookup.items() if s == sex}
        results = []
        for c in target_countries:
            per_code = {code: vals.get(c, 0.0) for code, vals in sex_codes.items()}
            results.append(calc_fn(per_code))
        return results

    for entry in template:
        kind = entry[0]

        if kind == 'blank':
            output_rows.append([None] * (7 + n_countries))

        elif kind == 'header':
            _, col3, col4, col5 = entry
            row = [None, None, None, col3, col4, col5, None] + [None] * n_countries
            output_rows.append(row)

        elif kind == 'data':
            _, sex, ghe_code = entry
            labels = GHE_LABEL_MAP.get(ghe_code, (None, None, None, str(ghe_code)))
            col2, col3, col4, col5 = labels
            country_vals = extract_country_vals(sex, ghe_code)
            row = [sex, ghe_code, col2, col3, col4, col5, None] + country_vals
            output_rows.append(row)

        elif kind == 'calc':
            _, sex, label_col5, col3, col4, calc_fn = entry
            country_vals = calc_country_vals(sex, calc_fn)
            row = [sex, None, None, col3, col4, label_col5, None] + country_vals
            output_rows.append(row)

    return output_rows


def build_population_row(sex, pop_lookup, target_countries):
    """Build the Population ('000) row for a given sex."""
    pops = pop_lookup.get(sex, {})
    vals = [pops.get(c, None) for c in target_countries]
    return [sex, None, None, "Population ('000) (2)", None, None, None] + vals


# WRITE OUTPUT SHEET

def write_sheet(ws, raw_header_rows, target_countries,
                pop_rows, data_rows, year, sheet_label=None):
    """
    Write a single categorised sheet.

    raw_header_rows : first 9 rows from the raw file (list of lists)
    target_countries: list of country name strings
    pop_rows        : list of population row lists (one per sex section present)
    data_rows       : list of data row lists from build_output_rows()
    year            : string like "2019"
    """
    styles = build_style_palette()

    raw_h6   = raw_header_rows[6]
    raw_h7   = raw_header_rows[7]
    raw_h8   = raw_header_rows[8]

    country_iso = {}
    country_inc = {}
    for ci, cn in enumerate(raw_h6):
        if pd.notna(cn) and ci > 6:
            if ci < len(raw_h7): country_iso[str(cn).strip()] = raw_h7[ci]
            if ci < len(raw_h8): country_inc[str(cn).strip()] = raw_h8[ci]

    text_map = {
        0: str(raw_header_rows[0][3]) if pd.notna(raw_header_rows[0][3]) else "World Health Organization",
        1: str(raw_header_rows[1][3]) if pd.notna(raw_header_rows[1][3]) else "Department of Data and Analytics",
        2: str(raw_header_rows[2][3]) if pd.notna(raw_header_rows[2][3]) else "",
        4: "Estimated DALY ('000) by cause, sex",
        5: f"and country or area (1), {year}",
    }
    for ri in range(6):
        txt = text_map.get(ri, "")
        if txt:
            cell = ws.cell(row=ri+1, column=4, value=txt)
            apply_cell_style(cell, styles['title'])

    ws.cell(row=7, column=1, value="Sex")
    ws.cell(row=7, column=2, value="GHE code")
    ws.cell(row=7, column=4, value="GHE cause")
    ws.cell(row=7, column=7, value="Country or area\n(See Notes for explanation of colour codes)")
    for ci, cn in enumerate(target_countries):
        ws.cell(row=7, column=8+ci, value=cn)
    for cell in ws[7][:7 + len(target_countries)]:
        apply_cell_style(cell, styles['header'])

    ws.cell(row=8, column=7, value="ISO-3 Code")
    for ci, cn in enumerate(target_countries):
        ws.cell(row=8, column=8+ci, value=country_iso.get(cn, ""))
    for cell in ws[8][:7 + len(target_countries)]:
        apply_cell_style(cell, styles['subheader'])

    ws.cell(row=9, column=7, value="Income group")
    for ci, cn in enumerate(target_countries):
        ws.cell(row=9, column=8+ci, value=country_inc.get(cn, ""))
    for cell in ws[9][:7 + len(target_countries)]:
        apply_cell_style(cell, styles['subheader'])

    current_row = 10
    pop_idx = 0

    sex_order = []
    seen = set()
    for dr in data_rows:
        s = dr[0]
        if s is not None and s not in seen:
            seen.add(s)
            sex_order.append(s)

    final_rows = []
    sex_ptr = None
    for dr in data_rows:
        s = dr[0]
        if s is not None and s != sex_ptr:
            sex_ptr = s
            if s in pop_lookup_global:
                final_rows.append(build_population_row(s, pop_lookup_global, target_countries))
        final_rows.append(dr)

    progress_stride = get_progress_stride(len(final_rows))

    for row_idx, dr in enumerate(final_rows, start=1):
        n_meta = 7
        ws.cell(row=current_row, column=1, value=dr[0])
        ws.cell(row=current_row, column=2, value=dr[1])
        ws.cell(row=current_row, column=3, value=dr[2])
        ws.cell(row=current_row, column=4, value=dr[3])
        ws.cell(row=current_row, column=5, value=dr[4])
        ws.cell(row=current_row, column=6, value=dr[5])
        for ci, val in enumerate(dr[7:]):
            ws.cell(row=current_row, column=8+ci, value=val)

        is_banded = current_row % 2 == 0
        style_data_row(ws, current_row, dr, len(target_countries), styles, is_banded=is_banded)
        if sheet_label and (row_idx == 1 or row_idx == len(final_rows) or row_idx % progress_stride == 0):
            print_progress(f"Writing {sheet_label}", row_idx, len(final_rows))
        current_row += 1

    ws.freeze_panes = 'H10'
    ws.sheet_view.showGridLines = False

    for merge_range in ('D1:F1', 'D2:F2', 'D3:F3', 'D5:F5', 'D6:F6'):
        ws.merge_cells(merge_range)

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 38
    ws.column_dimensions['F'].width = 42
    ws.column_dimensions['G'].width = 20
    for ci in range(len(target_countries)):
        col_letter = get_column_letter(8+ci)
        ws.column_dimensions[col_letter].width = 14

    for row_idx in range(1, current_row):
        if row_idx <= 6:
            ws.row_dimensions[row_idx].height = 20
        elif row_idx == 7:
            ws.row_dimensions[row_idx].height = 34
        elif row_idx in (8, 9):
            ws.row_dimensions[row_idx].height = 24


# MPNC HELPERS

def _mpnc_get_vals(ghe_code, data_lookup, target_countries):
    """
    Return a list of per-country float values for one MPNc GHE code.

    GHE 490 and 540 are computed from their sub-codes; all others are
    read directly from data_lookup[(sex='Persons', ghe_code)].
    """
    if ghe_code == 490:
        sub_codes = [500, 510, 520, 530]
        return [
            sum(data_lookup.get(("Persons", sc), {}).get(c, 0.0) for sc in sub_codes)
            for c in target_countries
        ]
    elif ghe_code == 540:
        sub_codes = [550, 560, 570, 580, 590]
        return [
            sum(data_lookup.get(("Persons", sc), {}).get(c, 0.0) for sc in sub_codes)
            for c in target_countries
        ]
    else:
        row = data_lookup.get(("Persons", ghe_code), {})
        return [row.get(c, 0.0) for c in target_countries]


def _mpnc_total_vals(data_lookup, target_countries):
    """
    Compute Total MPNc per country.

    Total MPNc = GHE 420 (Maternal)
               + GHE 490 computed (Neonatal)   = sum(500, 510, 520, 530)
               + GHE 540 computed (Nutritional) = sum(550, 560, 570, 580, 590)
               + GHE 1505 (SIDS)

    Individual sub-codes are used directly to avoid double-counting the
    490 / 540 aggregate rows.
    """
    leaf_codes = [420, 500, 510, 520, 530, 550, 560, 570, 580, 590, 1505]
    return [
        sum(data_lookup.get(("Persons", sc), {}).get(c, 0.0) for sc in leaf_codes)
        for c in target_countries
    ]


# CHART PLACEHOLDER

def write_chart_placeholder_sheet(ws, chart_title, data_sheet_hint):
    """
    Write a placeholder for a chart-only sheet.

    The original workbook stores these sheets as embedded Excel chart objects
    that openpyxl cannot reproduce programmatically.  This placeholder records
    the chart title and points to the source data sheet for manual recreation.
    """
    styles = build_style_palette()

    ws.cell(row=1, column=1, value=f"[CHART PLACEHOLDER: {chart_title}]")
    apply_cell_style(ws.cell(row=1, column=1), styles['title'])

    ws.cell(row=2, column=1, value=f"Source data tab: {data_sheet_hint}")
    apply_cell_style(ws.cell(row=2, column=1), styles['section'])

    ws.cell(row=3, column=1,
            value="To recreate: select the source data range in the indicated tab "
                  "and insert the appropriate chart type via Excel's Insert > Chart.")
    apply_cell_style(ws.cell(row=3, column=1), styles['body'])

    ws.column_dimensions['A'].width = 80
    ws.sheet_view.showGridLines = False


# MPNC DATA SHEET

def write_mpnc_data_sheet(ws, raw_header_rows, target_countries,
                          all_data_lookups, pop_lookup, year,
                          include_income_row=True, sheet_label=None):
    """
    Write an MPNc data sheet.

    Structure mirrors the source file's MPNc and MPNc (2) tabs:
      Rows 1-6  : WHO metadata header
      Row 7     : column headers  (Sex | GHE code | _ | Age Group | _ | GHE cause | _ | countries)
      Row 8     : ISO-3 codes
      Row 9     : income group codes  (include_income_row=True  → "MPNc" tab)
                  OR population row  (include_income_row=False → "MPNc (2)" tab)
      Row 10+   : age-group data blocks, each preceded by a blank separator row
                  Block structure: 1 "Total MPNc" summary row + 13 detail rows

    all_data_lookups : dict  { sheet_name -> data_lookup }  for all loaded sheets
    pop_lookup       : population lookup for the "All ages" sheet (Persons row)
    """
    styles = build_style_palette()
    n = len(target_countries)

    raw_h6 = raw_header_rows[6]
    raw_h7 = raw_header_rows[7]
    raw_h8 = raw_header_rows[8]
    country_iso = {}
    country_inc = {}
    for ci, cn in enumerate(raw_h6):
        if pd.notna(cn) and ci > 6:
            country_iso[str(cn).strip()] = raw_h7[ci] if ci < len(raw_h7) else ""
            country_inc[str(cn).strip()] = raw_h8[ci] if ci < len(raw_h8) else ""

    text_map = {
        0: (str(raw_header_rows[0][3]) if pd.notna(raw_header_rows[0][3])
            else "World Health Organization"),
        1: (str(raw_header_rows[1][3]) if pd.notna(raw_header_rows[1][3])
            else "Department of Data and Analytics"),
        2: (str(raw_header_rows[2][3]) if pd.notna(raw_header_rows[2][3]) else ""),
        4: "Estimated DALY ('000) by cause, sex",
        5: f"and country or area (1), {year}",
    }
    for ri in range(6):
        txt = text_map.get(ri, "")
        if txt:
            cell = ws.cell(row=ri + 1, column=4, value=txt)
            apply_cell_style(cell, styles['title'])

    ws.cell(row=7, column=1, value="Sex")
    ws.cell(row=7, column=2, value="GHE code")
    ws.cell(row=7, column=4, value="Age Group")
    ws.cell(row=7, column=6, value="GHE cause")
    for ci, cn in enumerate(target_countries):
        ws.cell(row=7, column=8 + ci, value=cn)
    for cell in ws[7][:7 + n]:
        apply_cell_style(cell, styles['header'])
    ws.row_dimensions[7].height = 34

    ws.cell(row=8, column=7, value="ISO-3 Code")
    for ci, cn in enumerate(target_countries):
        ws.cell(row=8, column=8 + ci, value=country_iso.get(cn, ""))
    for cell in ws[8][:7 + n]:
        apply_cell_style(cell, styles['subheader'])
    ws.row_dimensions[8].height = 24

    if include_income_row:
        ws.cell(row=9, column=7, value="Income group")
        for ci, cn in enumerate(target_countries):
            ws.cell(row=9, column=8 + ci, value=COUNTRY_INCOME_GROUPS.get(cn, ""))
        for cell in ws[9][:7 + n]:
            apply_cell_style(cell, styles['subheader'])
    else:
        ws.cell(row=9, column=1, value="Persons")
        ws.cell(row=9, column=4, value="Population ('000)")
        pops = pop_lookup.get("Persons", {})
        for ci, cn in enumerate(target_countries):
            cell = ws.cell(row=9, column=8 + ci, value=pops.get(cn, None))
            apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
        for col in range(1, 8):
            apply_cell_style(ws.cell(row=9, column=col), styles['population'])
    ws.row_dimensions[9].height = 24

    current_row = 10 if include_income_row else 11

    for block_idx, (age_label, sheet_name) in enumerate(MPNC_AGE_GROUPS):
        data_lkp = all_data_lookups.get(sheet_name, {})

        if block_idx > 0 or not include_income_row:
            ws.row_dimensions[current_row].height = 8
            current_row += 1

        totals = _mpnc_total_vals(data_lkp, target_countries)
        ws.cell(row=current_row, column=4, value=age_label)
        ws.cell(row=current_row, column=6, value="Total MPNc")
        for ci, v in enumerate(totals):
            cell = ws.cell(row=current_row, column=8 + ci, value=v)
            apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
        for col in range(1, 8):
            apply_cell_style(ws.cell(row=current_row, column=col), styles['summary'])
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for ghe_code in MPNC_DISPLAY_ORDER:
            label = MPNC_GHE_LABELS.get(ghe_code, str(ghe_code))
            vals = _mpnc_get_vals(ghe_code, data_lkp, target_countries)

            ws.cell(row=current_row, column=1, value="Persons")
            ws.cell(row=current_row, column=2, value=ghe_code)
            ws.cell(row=current_row, column=4, value=age_label)
            ws.cell(row=current_row, column=6, value=label)
            for ci, v in enumerate(vals):
                cell = ws.cell(row=current_row, column=8 + ci, value=v)
                apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)

            style_key = 'calc' if ghe_code in (490, 540) else 'body'
            for col in range(1, 8):
                apply_cell_style(ws.cell(row=current_row, column=col), styles[style_key])
            ws.row_dimensions[current_row].height = 18
            current_row += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 15
    for ci in range(n):
        ws.column_dimensions[get_column_letter(8 + ci)].width = 14

    for ri in range(1, 7):
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = 'H10'
    ws.sheet_view.showGridLines = False

    if sheet_label:
        print(f"    {sheet_label}: written {(len(MPNC_AGE_GROUPS) * (1 + len(MPNC_DISPLAY_ORDER))):,} data rows")


# MPNC CALCULATED SHEET

def write_mpnc_calculated_sheet(ws, all_data_lookups, pop_lookup,
                                focused_countries, year):
    """
    Write the 'MPNc Calculated' summary/statistics sheet.

    Layout (all 'All ages' Persons data):
      Row 1   : country names header (focused countries + "Global" placeholder)
      Row 2   : ISO-3 codes
      Row 3   : Total Population Size ('000)
      Row 4   : Total DALYs  (GHE 0, Persons)
      Row 5   : % of Total DALYs attributable to MPNc  (=Total MPNc / Total DALYs)
      Rows 6-18: 13 MPNc component rows (420, 490*, 500..530, 540*, 550..590, 1505)
                 * = calculated aggregate row
      Row 19-20: blank spacers
      Row 21  : "Chart Data" section label
      Row 22  : chart header: Country | MPNc % | Other DALYs %
      Rows 23+: one row per focused country  +  Global placeholder row

    NOTE: The "Global" column values in the source file come from WHO's
    worldwide aggregate, which is not available in the regional raw file.
    Those cells are left empty; fill them manually from the WHO GHE global
    estimates if needed.
    """
    styles = build_style_palette()
    all_lkp = all_data_lookups.get("All ages", {})
    n = len(focused_countries)
    global_col = 2 + n

    ws.cell(row=1, column=1, value=None)
    for ci, cn in enumerate(focused_countries):
        ws.cell(row=1, column=2 + ci, value=cn)
    ws.cell(row=1, column=global_col, value="Global ")
    for cell in ws[1][:global_col]:
        apply_cell_style(cell, styles['header'])
    ws.row_dimensions[1].height = 34

    for ci, cn in enumerate(focused_countries):
        ws.cell(row=2, column=2 + ci, value=MPNC_FOCUSED_ISO.get(cn, ""))
    for cell in ws[2][:global_col]:
        apply_cell_style(cell, styles['subheader'])
    ws.row_dimensions[2].height = 24

    ws.cell(row=3, column=1, value="Total Population Size ('000)")
    pops = pop_lookup.get("Persons", {})
    for ci, cn in enumerate(focused_countries):
        cell = ws.cell(row=3, column=2 + ci, value=pops.get(cn, None))
        apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
    apply_cell_style(ws.cell(row=3, column=1), styles['population'])
    ws.row_dimensions[3].height = 20

    ws.cell(row=4, column=1, value="Total DALYs")
    total_daly_row = all_lkp.get(("Persons", 0), {})
    for ci, cn in enumerate(focused_countries):
        cell = ws.cell(row=4, column=2 + ci, value=total_daly_row.get(cn, 0.0))
        apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
    apply_cell_style(ws.cell(row=4, column=1), styles['summary'])
    ws.row_dimensions[4].height = 20

    ws.cell(row=5, column=1, value="% of Total DALYs attributable to MPNc")
    mpnc_totals = _mpnc_total_vals(all_lkp, focused_countries)
    for ci, cn in enumerate(focused_countries):
        daly = total_daly_row.get(cn, 0.0)
        pct = (mpnc_totals[ci] / daly) if daly else 0.0
        cell = ws.cell(row=5, column=2 + ci, value=pct)
        cell.number_format = '0.00%'
        apply_cell_style(cell, styles['calc'])
    apply_cell_style(ws.cell(row=5, column=1), styles['calc'])
    ws.row_dimensions[5].height = 20

    current_row = 6
    for ghe_code in MPNC_DISPLAY_ORDER:
        label = MPNC_GHE_LABELS.get(ghe_code, str(ghe_code))
        ws.cell(row=current_row, column=1, value=label)
        vals = _mpnc_get_vals(ghe_code, all_lkp, focused_countries)
        for ci, v in enumerate(vals):
            cell = ws.cell(row=current_row, column=2 + ci, value=v)
            apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
        style_key = 'calc' if ghe_code in (490, 540) else 'body'
        apply_cell_style(ws.cell(row=current_row, column=1), styles[style_key])
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    for r in (19, 20):
        ws.row_dimensions[r].height = 8

    ws.cell(row=21, column=1, value="Chart Data")
    apply_cell_style(ws.cell(row=21, column=1), styles['section'])
    ws.row_dimensions[21].height = 22

    ws.cell(row=22, column=1, value="Country")
    ws.cell(row=22, column=2, value="MPNc %")
    ws.cell(row=22, column=3, value="Other DALYs %")
    for cell in ws[22][:3]:
        apply_cell_style(cell, styles['subheader'])
    ws.row_dimensions[22].height = 24

    chart_row = 23
    for ci, cn in enumerate(focused_countries):
        daly = total_daly_row.get(cn, 0.0)
        mpnc_pct = (mpnc_totals[ci] / daly) if daly else 0.0
        other_pct = 1.0 - mpnc_pct
        display_name = "Lao PDR" if "Lao People" in cn else cn
        ws.cell(row=chart_row, column=1, value=display_name)
        c_m = ws.cell(row=chart_row, column=2, value=mpnc_pct)
        c_o = ws.cell(row=chart_row, column=3, value=other_pct)
        c_m.number_format = '0.00%'
        c_o.number_format = '0.00%'
        for col in range(1, 4):
            apply_cell_style(ws.cell(row=chart_row, column=col), styles['body'])
        ws.row_dimensions[chart_row].height = 18
        chart_row += 1

    ws.cell(row=chart_row, column=1, value="Global")
    ws.cell(row=chart_row, column=2, value=None)
    ws.cell(row=chart_row, column=3, value=None)
    for col in range(1, 4):
        apply_cell_style(ws.cell(row=chart_row, column=col), styles['body'])
    ws.row_dimensions[chart_row].height = 18

    ws.column_dimensions['A'].width = 45
    for ci in range(n + 1):
        ws.column_dimensions[get_column_letter(2 + ci)].width = 16
    ws.sheet_view.showGridLines = False


# MAIN AGE GROUP SHEET

def write_main_age_group_sheet(ws, all_age_data_lookups, target_countries, year):
    """
    Write the 'MainAgeGroup' sheet, which contains two linked summary tables.

    ┌──────────────────────────────────────────────────────────────────┐
    │ TABLE 1  (rows 1-15)  –  Age × GHE cause cross-tab              │
    │  Columns : Sex | Age label | GHE cause | All Countries | [14]   │
    │  Causes  : GHE 20  (Infectious & parasitic diseases)             │
    │            GHE 600 (Noncommunicable diseases)                    │
    │  Ages    : 0-4, 5-14, 15-29, 30-49, 50-59, 60-69, 70+          │
    │  "All Countries" = rounded sum across all target countries       │
    └──────────────────────────────────────────────────────────────────┘
    ┌──────────────────────────────────────────────────────────────────┐
    │ TABLE 2  (rows 19+)  –  All-ages Persons: selected categories   │
    │  Column layout: Sex | GHE code | _ | col4 | col5 | col6 | _ |  │
    │                 [countries]                                       │
    │  D. Injuries    : 1510 (summary), 1520, 1630, 1620               │
    │  E. Mental      : section total, 1610, 830, 880, 890,            │
    │                   910* (=911+912), 911, 912                       │
    │  F. Substance   : section total, 860, 870                        │
    │  * GHE 910 is computed as sum(911, 912)                           │
    └──────────────────────────────────────────────────────────────────┘
    ┌──────────────────────────────────────────────────────────────────┐
    │ TABLE 3  (rows 39-40)  –  Chart support lookup row               │
    │  Single data row for GHE 1510 Injuries across all countries      │
    └──────────────────────────────────────────────────────────────────┘
    """
    styles = build_style_palette()
    all_lkp = all_age_data_lookups.get("All ages", {})
    n = len(target_countries)

    age_sheet_map = {
        "0-4yo":   "0-4",
        "5-14yo":  "5-14",
        "15-29yo": "15-29",
        "30-49yo": "30-49",
        "50-59yo": "50-59",
        "60-69yo": "60-69",
        "70+ yo":  "70+",
    }
    age_labels_ordered = ["0-4yo", "5-14yo", "15-29yo",
                          "30-49yo", "50-59yo", "60-69yo", "70+ yo"]

    table1_causes = [
        (20,  "Infectious and parasitic diseases"),
        (600, "Noncommunicable diseases"),
    ]

    ws.cell(row=1, column=1, value="Sex")
    ws.cell(row=1, column=2, value="Age")
    ws.cell(row=1, column=3, value="GHE cause")
    ws.cell(row=1, column=4, value="All Countries")
    for ci, cn in enumerate(target_countries):
        ws.cell(row=1, column=5 + ci, value=cn)
    for cell in ws[1][:4 + n]:
        apply_cell_style(cell, styles['header'])
    ws.row_dimensions[1].height = 34

    current_row = 2
    for ghe_code, cause_label in table1_causes:
        for age_lbl in age_labels_ordered:
            sheet_name = age_sheet_map[age_lbl]
            data_lkp = all_age_data_lookups.get(sheet_name, {})
            vals_dict = data_lkp.get(("Persons", ghe_code), {})
            vals = [vals_dict.get(c, 0.0) for c in target_countries]
            all_countries_sum = round(sum(vals), 1)

            ws.cell(row=current_row, column=1, value="Persons")
            ws.cell(row=current_row, column=2, value=age_lbl)
            ws.cell(row=current_row, column=3, value=cause_label)
            ws.cell(row=current_row, column=4, value=all_countries_sum)
            for ci, v in enumerate(vals):
                cell = ws.cell(row=current_row, column=5 + ci, value=v)
                apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
            is_banded = (current_row % 2 == 0)
            for col in range(1, 5):
                c = ws.cell(row=current_row, column=col)
                apply_cell_style(c, styles['body'])
                if is_banded:
                    c.fill = styles['band_fill']
            ws.row_dimensions[current_row].height = 18
            current_row += 1

    for r in range(current_row, 19):
        ws.row_dimensions[r].height = 8
    current_row = 19

    ws.cell(row=19, column=1, value="Sex")
    ws.cell(row=19, column=2, value="GHE code")
    ws.cell(row=19, column=4, value="GHE cause")
    ws.cell(row=19, column=6, value="GHE cause")
    for ci, cn in enumerate(target_countries):
        ws.cell(row=19, column=8 + ci, value=cn)
    for cell in ws[19][:7 + n]:
        apply_cell_style(cell, styles['header'])
    ws.row_dimensions[19].height = 34
    current_row = 20

    injuries_rows = [
        (1510, "D.", "Injuries incl War",                        True),
        (1520, None, "Unintentional injuries",                   False),
        (1630, None, "Collective violence and legal intervention",False),
        (1620, None, "Interpersonal violence",                   False),
    ]
    for ghe_code, col4_val, cause_label, is_summary_row in injuries_rows:
        vals_dict = all_lkp.get(("Persons", ghe_code), {})
        ws.cell(row=current_row, column=1, value="Persons")
        ws.cell(row=current_row, column=2, value=ghe_code)
        if is_summary_row:
            ws.cell(row=current_row, column=4, value=col4_val)
            ws.cell(row=current_row, column=5, value=cause_label)
        else:
            ws.cell(row=current_row, column=6, value=cause_label)
        for ci, cn in enumerate(target_countries):
            cell = ws.cell(row=current_row, column=8 + ci, value=vals_dict.get(cn, 0.0))
            apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
        style_key = 'summary' if is_summary_row else 'body'
        for col in range(1, 8):
            apply_cell_style(ws.cell(row=current_row, column=col), styles[style_key])
        ws.row_dimensions[current_row].height = 20 if is_summary_row else 18
        current_row += 1

    ws.row_dimensions[current_row].height = 8
    current_row += 1

    mh_leaf_codes = [1610, 830, 880, 890, 911, 912]
    mh_totals = [
        sum(all_lkp.get(("Persons", mhc), {}).get(cn, 0.0) for mhc in mh_leaf_codes)
        for cn in target_countries
    ]
    ws.cell(row=current_row, column=4, value="E.")
    ws.cell(row=current_row, column=5, value="Mental Disorders & Wellbeing")
    for ci, v in enumerate(mh_totals):
        cell = ws.cell(row=current_row, column=8 + ci, value=v)
        apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
    for col in range(1, 8):
        apply_cell_style(ws.cell(row=current_row, column=col), styles['section'])
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    mh_detail_rows = [
        (1610, "Self-harm",                                  False),
        (830,  "Depressive Disorders",                       False),
        (880,  "Anxiety disorders",                          False),
        (890,  "Eating disorders",                           False),
        (910,  "Childhood behavioural disorders",            True),
        (911,  "Attention deficit/hyperactivity syndrome",   False),
        (912,  "Conduct disorder",                           False),
    ]
    for ghe_code, cause_label, is_computed in mh_detail_rows:
        ws.cell(row=current_row, column=1, value="Persons")
        ws.cell(row=current_row, column=2, value=ghe_code)
        ws.cell(row=current_row, column=6, value=cause_label)
        if is_computed:
            for ci, cn in enumerate(target_countries):
                v = (all_lkp.get(("Persons", 911), {}).get(cn, 0.0) +
                     all_lkp.get(("Persons", 912), {}).get(cn, 0.0))
                cell = ws.cell(row=current_row, column=8 + ci, value=v)
                apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
            style_key = 'calc'
        else:
            vals_dict = all_lkp.get(("Persons", ghe_code), {})
            for ci, cn in enumerate(target_countries):
                cell = ws.cell(row=current_row, column=8 + ci, value=vals_dict.get(cn, 0.0))
                apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
            style_key = 'body'
        for col in range(1, 8):
            apply_cell_style(ws.cell(row=current_row, column=col), styles[style_key])
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    ws.row_dimensions[current_row].height = 8
    current_row += 1

    sa_codes = [860, 870]
    sa_totals = [
        sum(all_lkp.get(("Persons", sc), {}).get(cn, 0.0) for sc in sa_codes)
        for cn in target_countries
    ]
    ws.cell(row=current_row, column=4, value="F.")
    ws.cell(row=current_row, column=5, value="Substance Abuse")
    for ci, v in enumerate(sa_totals):
        cell = ws.cell(row=current_row, column=8 + ci, value=v)
        apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
    for col in range(1, 8):
        apply_cell_style(ws.cell(row=current_row, column=col), styles['section'])
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    sa_detail_rows = [(860, "Alcohol use disorders"), (870, "Drug use disorders")]
    for ghe_code, cause_label in sa_detail_rows:
        ws.cell(row=current_row, column=1, value="Persons")
        ws.cell(row=current_row, column=2, value=ghe_code)
        ws.cell(row=current_row, column=6, value=cause_label)
        vals_dict = all_lkp.get(("Persons", ghe_code), {})
        for ci, cn in enumerate(target_countries):
            cell = ws.cell(row=current_row, column=8 + ci, value=vals_dict.get(cn, 0.0))
            apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
        for col in range(1, 8):
            apply_cell_style(ws.cell(row=current_row, column=col), styles['body'])
        ws.row_dimensions[current_row].height = 18
        current_row += 1

    chart_support_row = 39
    ws.cell(row=chart_support_row, column=1, value="GHE Causes")
    for ci, cn in enumerate(target_countries):
        ws.cell(row=chart_support_row, column=2 + ci, value=cn)
    for cell in ws[chart_support_row][:1 + n]:
        apply_cell_style(cell, styles['subheader'])
    ws.row_dimensions[chart_support_row].height = 22

    inj_vals = all_lkp.get(("Persons", 1510), {})
    data_row = chart_support_row + 1
    for ci, cn in enumerate(target_countries):
        cell = ws.cell(row=data_row, column=2 + ci, value=inj_vals.get(cn, 0.0))
        apply_cell_style(cell, styles['numeric'], NUMBER_FORMAT)
    ws.row_dimensions[data_row].height = 18

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 38
    ws.column_dimensions['G'].width = 5
    for ci in range(n):
        ws.column_dimensions[get_column_letter(8 + ci)].width = 14
    for ci in range(n):
        col_letter_t1 = get_column_letter(5 + ci)
        if ws.column_dimensions[col_letter_t1].width < 14:
            ws.column_dimensions[col_letter_t1].width = 14

    ws.freeze_panes = 'E2'
    ws.sheet_view.showGridLines = False


# MAIN

pop_lookup_global = {}


def main():
    global pop_lookup_global

    print(f"Loading raw file: {RAW_FILE}")
    if not os.path.exists(RAW_FILE):
        raise FileNotFoundError(f"Raw file not found: {RAW_FILE}")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    total_output_sheets = 17
    completed_sheets = 0

    print(f"  Processing sheet {completed_sheets + 1}/{total_output_sheets}: All ages ORIGINAL")
    raw_header_rows, country_cols, data_lkp_all, pop_lkp_all = \
        load_raw_sheet(RAW_FILE, 'All ages')

    ws_orig = wb_out.create_sheet("All ages ORIGINAL")
    wb_raw = load_workbook(RAW_FILE, read_only=True, data_only=True)
    ws_raw = wb_raw['All ages']
    total_raw_rows = ws_raw.max_row or 0
    copy_stride = get_progress_stride(total_raw_rows)
    for row_idx, row in enumerate(ws_raw.iter_rows(values_only=True), start=1):
        ws_orig.append(list(row))
        if row_idx == 1 or row_idx == total_raw_rows or row_idx % copy_stride == 0:
            print_progress("Copying All ages ORIGINAL", row_idx, total_raw_rows)
    wb_raw.close()
    format_full_sheet(ws_orig, progress_label="Formatting All ages ORIGINAL")
    completed_sheets += 1
    print(f"  Completed sheet {completed_sheets}/{total_output_sheets}: All ages ORIGINAL")

    print(f"  Processing sheet {completed_sheets + 1}/{total_output_sheets}: All ages Persons Categorised")
    pop_lookup_global = pop_lkp_all
    ws_pc = wb_out.create_sheet("All ages Persons Categorised")
    template_pc = _persons_all_ages_template()
    data_rows_pc = build_output_rows(template_pc, data_lkp_all, pop_lkp_all, TARGET_COUNTRIES)
    write_sheet(ws_pc, raw_header_rows, TARGET_COUNTRIES,
                [], data_rows_pc, YEAR, sheet_label="All ages Persons Categorised")
    completed_sheets += 1
    print(f"  Completed sheet {completed_sheets}/{total_output_sheets}: All ages Persons Categorised")

    print(f"  Processing sheet {completed_sheets + 1}/{total_output_sheets}: All ages Gender Categorised")
    ws_gc = wb_out.create_sheet("All ages Gender Categorised")
    template_gc = []
    for sex in ('Males', 'Females'):
        template_gc.extend(_sex_section_template(sex))
        template_gc.append(('blank',))
    data_rows_gc = build_output_rows(template_gc, data_lkp_all, pop_lkp_all, TARGET_COUNTRIES)
    write_sheet(ws_gc, raw_header_rows, TARGET_COUNTRIES,
                [], data_rows_gc, YEAR, sheet_label="All ages Gender Categorised")
    completed_sheets += 1
    print(f"  Completed sheet {completed_sheets}/{total_output_sheets}: All ages Gender Categorised")

    all_age_data_lookups = {"All ages": data_lkp_all}

    age_sheets = ['0-4', '5-14', '15-29', '30-49', '50-59', '60-69', '70+']
    for age in age_sheets:
        print(f"  Processing sheet {completed_sheets + 1}/{total_output_sheets}: Categorised({age})")
        raw_h, _, data_lkp, pop_lkp = load_raw_sheet(RAW_FILE, age)
        all_age_data_lookups[age] = data_lkp
        pop_lookup_global = pop_lkp
        ws_age = wb_out.create_sheet(f"Categorised({age})")
        template_age = _age_group_template()
        data_rows_age = build_output_rows(template_age, data_lkp, pop_lkp, TARGET_COUNTRIES)
        write_sheet(ws_age, raw_h, TARGET_COUNTRIES,
                    [], data_rows_age, YEAR, sheet_label=f"Categorised({age})")
        completed_sheets += 1
        print(f"  Completed sheet {completed_sheets}/{total_output_sheets}: Categorised({age})")

    print(f"  Processing sheet {completed_sheets + 1}/{total_output_sheets}: Chart – Infectious Diseases")
    ws_ch1 = wb_out.create_sheet("Chart – Infectious Diseases")
    write_chart_placeholder_sheet(
        ws_ch1,
        chart_title="Infectious Diseases burden by country",
        data_sheet_hint="All ages Persons Categorised  (rows for GHE 20 and sub-codes)",
    )
    completed_sheets += 1
    print(f"  Completed sheet {completed_sheets}/{total_output_sheets}: Chart – Infectious Diseases")

    print(f"  Processing sheet {completed_sheets + 1}/{total_output_sheets}: MPNc Calculated")
    ws_mpnc_calc = wb_out.create_sheet("MPNc Calculated")
    write_mpnc_calculated_sheet(
        ws_mpnc_calc,
        all_data_lookups=all_age_data_lookups,
        pop_lookup=pop_lkp_all,
        focused_countries=MPNC_FOCUSED_COUNTRIES,
        year=YEAR,
    )
    completed_sheets += 1
    print(f"  Completed sheet {completed_sheets}/{total_output_sheets}: MPNc Calculated")

    print(f"  Processing sheet {completed_sheets + 1}/{total_output_sheets}: Chart – MPNc (All Countries)")
    ws_ch2 = wb_out.create_sheet("Chart – MPNc (All Countries)")
    write_chart_placeholder_sheet(
        ws_ch2,
        chart_title="MPNc burden – All 14 target countries",
        data_sheet_hint="MPNc  (Total MPNc row for the 'All' age group)",
    )
    completed_sheets += 1
    print(f"  Completed sheet {completed_sheets}/{total_output_sheets}: Chart – MPNc (All Countries)")

    print(f"  Processing sheet {completed_sheets + 1}/{total_output_sheets}: MPNc (2)")
    ws_mpnc2 = wb_out.create_sheet("MPNc (2)")
    write_mpnc_data_sheet(
        ws_mpnc2,
        raw_header_rows=raw_header_rows,
        target_countries=MPNC_FOCUSED_COUNTRIES,
        all_data_lookups=all_age_data_lookups,
        pop_lookup=pop_lkp_all,
        year=YEAR,
        include_income_row=False,
        sheet_label="MPNc (2)",
    )
    completed_sheets += 1
    print(f"  Completed sheet {completed_sheets}/{total_output_sheets}: MPNc (2)")

    print(f"  Processing sheet {completed_sheets + 1}/{total_output_sheets}: MPNc")
    ws_mpnc = wb_out.create_sheet("MPNc")
    write_mpnc_data_sheet(
        ws_mpnc,
        raw_header_rows=raw_header_rows,
        target_countries=TARGET_COUNTRIES,
        all_data_lookups=all_age_data_lookups,
        pop_lookup=pop_lkp_all,
        year=YEAR,
        include_income_row=True,
        sheet_label="MPNc",
    )
    completed_sheets += 1
    print(f"  Completed sheet {completed_sheets}/{total_output_sheets}: MPNc")

    print(f"  Processing sheet {completed_sheets + 1}/{total_output_sheets}: Chart – MPNc (Focused)")
    ws_ch3 = wb_out.create_sheet("Chart – MPNc (Focused)")
    write_chart_placeholder_sheet(
        ws_ch3,
        chart_title="MPNc burden – focused country subset (stacked bar: MPNc% vs Other%)",
        data_sheet_hint="MPNc Calculated  (Chart Data section, rows 21–30)",
    )
    completed_sheets += 1
    print(f"  Completed sheet {completed_sheets}/{total_output_sheets}: Chart – MPNc (Focused)")

    print(f"  Processing sheet {completed_sheets + 1}/{total_output_sheets}: MainAgeGroup")
    ws_mag = wb_out.create_sheet("MainAgeGroup")
    write_main_age_group_sheet(
        ws_mag,
        all_age_data_lookups=all_age_data_lookups,
        target_countries=TARGET_COUNTRIES,
        year=YEAR,
    )
    completed_sheets += 1
    print(f"  Completed sheet {completed_sheets}/{total_output_sheets}: MainAgeGroup")

    print(f"\nSaving workbook to {OUTPUT_FILE} ... this can pause briefly on large files")
    wb_out.save(OUTPUT_FILE)
    print(f"\nDone! Output saved to: {OUTPUT_FILE}")
    print(f"Total sheets written: {completed_sheets}")


if __name__ == "__main__":
    main()
